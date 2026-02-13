from flask import make_response
import pandas as pd
import subprocess
from io import BytesIO
from config import *
import pdfkit
import os
import random
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ManipulacaoArquivos:
    """
    Classe responsável devolver um arquivo PDF como resposta a uma
    requisição.
    """
    
    def gerar_pdf_from_html(html, nome_arquivo, orientacao="Portrait", abrir_em_nova_aba=True):
        """
        Recebe como parâmetro um template HTML renderizado, um nome para o arquivo
        PDF de saída, a orientação do documento e se deve abrir em nova aba.
        """
        options = {
            "enable-local-file-access": "",  # habilita acesso a arquivos locais
            "encoding": "UTF-8",
            "orientation": orientacao,  # Portrait ou Landscape
        }
        
        # Gerar o PDF a partir do HTML
        arquivo_pdf = pdfkit.from_string(
            html, False, configuration=pdfkit_config, options=options
        )
        
        # Retorna o PDF como resposta da requisição
        resposta = make_response(arquivo_pdf)
        resposta.headers["Content-Type"] = "application/pdf"
        
        # Define se vai baixar ou abrir em nova aba
        if abrir_em_nova_aba:
            resposta.headers["Content-Disposition"] = f"inline; filename={nome_arquivo}.pdf"
        else:
            resposta.headers["Content-Disposition"] = f"attachment; filename={nome_arquivo}.pdf"
        
        return resposta
    
    @staticmethod
    def gerar_imagem_from_html(html, nome_arquivo, largura=1400, altura=0):
        import subprocess
        from flask import make_response
        import os
        
        wkhtmltopdf_path = caminho_wkhtmltopdf

        if os.name == 'nt':  # Windows
            wkhtmltoimage_path = wkhtmltopdf_path.replace('wkhtmltopdf.exe', 'wkhtmltoimage.exe')
        else:  # Linux ou outros
            wkhtmltoimage_path = wkhtmltopdf_path.replace('wkhtmltopdf', 'wkhtmltoimage')
                
        # Verificar se o arquivo existe
        if not os.path.exists(wkhtmltoimage_path):
            raise Exception(f"wkhtmltoimage não encontrado em: {wkhtmltoimage_path}")
        
        options = [
            '--width', str(largura),
            '--height', str(altura),
            '--format', 'jpg',
            '--quality', '95',
            '--enable-local-file-access',
            '--encoding', 'UTF-8'
        ]
        
        try:
            processo = subprocess.run([
                wkhtmltoimage_path,
                *options,
                '-',
                '-'
            ], input=html.encode('utf-8'), capture_output=True, timeout=30)
            
            if processo.returncode != 0:
                raise Exception(f"Erro ao gerar imagem: {processo.stderr.decode()}")
            
            resposta = make_response(processo.stdout)
            resposta.headers['Content-Type'] = 'image/jpeg'
            resposta.headers['Content-Disposition'] = f'inline; filename={nome_arquivo}.jpg'
            
            return resposta
            
        except Exception as e:
            raise Exception(f"Erro ao gerar imagem: {str(e)}")

    def mover_e_renomear_arquivo(caminho_atual, novo_caminho, novo_nome=None):
        """
        Recebe como parâmetro o 'caminho atual' do arquivo (diretorio/nome.extensao),
        o 'novo caminho' (diretório/) e o 'novo nome' para arquivo (novo_nome_arquivo.extensao).
        A função verifica se o 'caminho_atual' e o 'novo caminho' existem e se o 'nome do arquivo'
        foi enviado. Caso sim, move e renomeia o arquivo e retorna um dicionário com a chave
        'validado' e a mensagem de sucesso. Caso não, retorna um dicionário com chave 'erro'
        e mensagem de erro. OBS: A função adiciona um número aleatório ao novo nome do arquivo
        para evitar conflitos de nomes iguais.
        """
        resultado = {}
        if (
            os.path.exists(caminho_atual)
            and os.path.exists(novo_caminho)
            and novo_nome != None
        ):
            novo = novo_caminho + str(random.randint(1, 99)) + "_" + novo_nome
            os.rename(caminho_atual, novo)
            resultado["validado"] = f"Arquivo movido e renomeado!"
            
            return resultado
        
        else:
            resultado["erro"] = f"Operação de mover e renomear falhou"
            return resultado

    def exportar_excel(dados, nome_arquivo, colunas=None):
        """
        Gera e retorna um arquivo Excel para download.
        
        :param dados: Pode ser uma lista de dicionários, uma lista de listas (com colunas), ou um DataFrame.
        :param nome_arquivo: Nome do arquivo (sem extensão).
        :param colunas: Lista de nomes das colunas (se dados for lista de listas). Opcional.
        """
        # Converte os dados para DataFrame, se já não for
        if isinstance(dados, pd.DataFrame):
            df = dados
        else:
            df = pd.DataFrame(dados, columns=colunas)
        
        # Cria um buffer para manter o arquivo na memória
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
        
        output.seek(0)
        
        resposta = make_response(output.read())
        resposta.headers["Content-Disposition"] = (
            f"attachment; filename={nome_arquivo}.xlsx"
        )
        resposta.headers["Content-Type"] = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        return resposta

    @staticmethod
    def exportar_excel_formatado(dados, nome_arquivo, titulo_planilha='Relatório',
                                  colunas_monetarias=None, coluna_destaque=None,
                                  linha_totais=None):
        """
        Gera um Excel formatado com:
        - Linha de título mesclada
        - Cabeçalho estilizado (verde escuro)
        - Largura de colunas automática
        - Colunas monetárias formatadas em BRL (R$ 1.234,56)
        - Linha de totais opcional ao final
        - Coluna de destaque (cor vermelha) para saldos pendentes

        :param dados: Lista de dicts (cada dict = 1 linha).
        :param nome_arquivo: Nome do arquivo sem extensão.
        :param titulo_planilha: Título exibido na primeira linha.
        :param colunas_monetarias: Lista de nomes de colunas que contêm valores monetários (float).
        :param coluna_destaque: Nome de coluna monetária a destacar em vermelho.
        :param linha_totais: Dict com labels de totalização (ex: {'Valor Original': 123456.78}).
        """
        colunas_monetarias = colunas_monetarias or []

        df = pd.DataFrame(dados)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Dados', startrow=2)
            wb = writer.book
            ws = writer.sheets['Dados']

            num_colunas = len(df.columns)
            num_linhas = len(df)

            # --- Estilos ---
            verde_escuro = '1E5631'
            verde_claro = 'E7F0E7'
            cinza_zebra = 'F9F9F9'
            vermelho = 'C92A2A'

            font_titulo = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
            fill_titulo = PatternFill(start_color=verde_escuro, end_color=verde_escuro, fill_type='solid')

            font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
            fill_header = PatternFill(start_color=verde_escuro, end_color=verde_escuro, fill_type='solid')

            font_normal = Font(name='Segoe UI', size=10, color='333333')
            font_monetario = Font(name='Segoe UI', size=10, color='333333')
            font_destaque = Font(name='Segoe UI', size=10, bold=True, color=vermelho)

            font_total = Font(name='Segoe UI', size=11, bold=True, color=verde_escuro)
            fill_total = PatternFill(start_color=verde_claro, end_color=verde_claro, fill_type='solid')

            fill_zebra = PatternFill(start_color=cinza_zebra, end_color=cinza_zebra, fill_type='solid')

            align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            align_right = Alignment(horizontal='right', vertical='center')
            align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

            borda_fina = Border(
                bottom=Side(style='thin', color='DDDDDD')
            )

            # --- Título (linha 1, mesclada) ---
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_colunas)
            celula_titulo = ws.cell(row=1, column=1, value=titulo_planilha)
            celula_titulo.font = font_titulo
            celula_titulo.fill = fill_titulo
            celula_titulo.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 35

            # --- Linha vazia (row 2) ---
            ws.row_dimensions[2].height = 8

            # --- Cabeçalho (row 3) ---
            for col_idx in range(1, num_colunas + 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = Border(bottom=Side(style='medium', color=verde_escuro))
            ws.row_dimensions[3].height = 28

            # --- Identificar colunas monetárias e destaque por índice ---
            colunas = list(df.columns)
            idx_monetarias = set()
            idx_destaque = None
            for i, col_name in enumerate(colunas):
                if col_name in colunas_monetarias:
                    idx_monetarias.add(i)
                if col_name == coluna_destaque:
                    idx_destaque = i

            # --- Dados (a partir da row 4) ---
            for row_idx in range(num_linhas):
                excel_row = row_idx + 4

                for col_idx in range(num_colunas):
                    cell = ws.cell(row=excel_row, column=col_idx + 1)
                    cell.border = borda_fina

                    if col_idx in idx_monetarias:
                        valor = cell.value
                        if isinstance(valor, (int, float)):
                            cell.number_format = '#,##0.00'
                            cell.value = round(valor, 2)
                        cell.alignment = align_center

                        if col_idx == idx_destaque:
                            cell.font = font_destaque
                        else:
                            cell.font = font_monetario
                    else:
                        cell.font = font_normal
                        cell.alignment = align_center

                # Zebra striping
                if row_idx % 2 == 1:
                    for col_idx in range(num_colunas):
                        ws.cell(row=excel_row, column=col_idx + 1).fill = fill_zebra

            # --- Linha de totais ---
            if linha_totais:
                total_row = num_linhas + 4
                ws.row_dimensions[total_row].height = 28

                for col_idx in range(num_colunas):
                    cell = ws.cell(row=total_row, column=col_idx + 1)
                    cell.fill = fill_total
                    cell.border = Border(top=Side(style='medium', color=verde_escuro))
                    col_name = colunas[col_idx]

                    if col_name in linha_totais:
                        cell.value = round(linha_totais[col_name], 2)
                        cell.number_format = '#,##0.00'
                        cell.alignment = align_center
                        if col_name == coluna_destaque:
                            cell.font = Font(name='Segoe UI', size=11, bold=True, color=vermelho)
                        else:
                            cell.font = font_total
                    elif col_idx == 0:
                        cell.value = 'TOTAIS'
                        cell.font = font_total
                        cell.alignment = align_center
                    else:
                        cell.font = font_total

            # --- Auto-width das colunas ---
            for col_idx in range(num_colunas):
                col_letter = get_column_letter(col_idx + 1)
                max_len = len(str(colunas[col_idx])) + 4

                for row_idx in range(num_linhas):
                    cell_value = ws.cell(row=row_idx + 4, column=col_idx + 1).value
                    if cell_value is not None:
                        cell_len = len(str(cell_value))
                        if cell_len > max_len:
                            max_len = cell_len

                ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, 45))

            # Congelar cabeçalho
            ws.freeze_panes = 'A4'

        output.seek(0)

        resposta = make_response(output.read())
        resposta.headers['Content-Disposition'] = f'attachment; filename={nome_arquivo}.xlsx'
        resposta.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        return resposta

    @staticmethod
    def exportar_excel_agrupado_ap_pendentes(grupos, nome_arquivo, titulo_planilha='AP - Pendentes'):
        """
        Gera um Excel com agrupamento visual por faturamento.
        Cada faturamento tem uma linha de título destacada, 
        seguida das cargas e uma linha de subtotais.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Dados'
        
        # Estilos
        verde_escuro = '1E5631'
        verde_claro = 'E8F5E8'
        azul_escuro = '1565C0'
        vermelho = 'C92A2A'
        cinza = 'F5F5F5'
        
        font_titulo = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
        fill_titulo = PatternFill(start_color=verde_escuro, end_color=verde_escuro, fill_type='solid')
        
        font_grupo = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        fill_grupo = PatternFill(start_color=verde_escuro, end_color=verde_escuro, fill_type='solid')
        
        font_header = Font(name='Segoe UI', size=10, bold=True, color=verde_escuro)
        fill_header = PatternFill(start_color=verde_claro, end_color=verde_claro, fill_type='solid')
        
        font_normal = Font(name='Segoe UI', size=10, color='333333')
        font_valor = Font(name='Segoe UI', size=10, bold=True, color=verde_escuro)
        font_pendente = Font(name='Segoe UI', size=10, bold=True, color=vermelho)
        
        font_subtotal = Font(name='Segoe UI', size=10, bold=True, color=verde_escuro)
        fill_subtotal = PatternFill(start_color=verde_claro, end_color=verde_claro, fill_type='solid')
        
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        
        borda_fina = Border(bottom=Side(style='thin', color='DDDDDD'))
        borda_grupo = Border(bottom=Side(style='medium', color=verde_escuro))
        
        # Colunas das cargas
        colunas_cargas = ['Data', 'Tipo', 'Entidade', 'Cliente', 'Produto', 'Bitola', 'NF', 'Peso (Ton)', 'Preço', 'Valor Final']
        num_colunas = len(colunas_cargas)
        
        # Título do relatório (linha 1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_colunas)
        celula_titulo = ws.cell(row=1, column=1, value=titulo_planilha)
        celula_titulo.font = font_titulo
        celula_titulo.fill = fill_titulo
        celula_titulo.alignment = align_center
        ws.row_dimensions[1].height = 30
        
        row = 3  # Começa na linha 3
        
        total_geral_saldo = 0
        
        for grupo in grupos:
            codigo_fat = grupo.get('codigo_faturamento', '-')
            pessoa_nome = grupo.get('pessoa_nome', '-')
            tipo_operacao = grupo.get('tipo_operacao', '-')
            data_vencimento = grupo.get('data_vencimento_mais_antiga')
            dias_atraso = grupo.get('maior_atraso', 0)
            saldo_pendente = (grupo.get('total_pendente_100') or 0) / 100
            valor_bruto_total = (grupo.get('valor_bruto_total') or 0) / 100
            total_original = (grupo.get('total_original_100') or 0) / 100
            data_faturamento = grupo.get('data_faturamento')
            
            total_geral_saldo += saldo_pendente
            
            # Status de atraso
            if dias_atraso > 0:
                status = f'{dias_atraso} dias atraso'
            elif dias_atraso == 0:
                status = 'Vence hoje'
            else:
                status = 'No prazo'
            
            # Linha do grupo (faturamento)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_colunas)
            info_grupo = f'{codigo_fat} - {pessoa_nome[:40]} | {tipo_operacao} | Venc: {data_vencimento.strftime("%d/%m/%Y") if data_vencimento else "-"} | {status} | Saldo: R$ {saldo_pendente:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            celula_grupo = ws.cell(row=row, column=1, value=info_grupo)
            celula_grupo.font = font_grupo
            celula_grupo.fill = fill_grupo
            celula_grupo.alignment = align_left
            ws.row_dimensions[row].height = 25
            row += 1
            
            detalhes = grupo.get('detalhes_cargas')
            
            if detalhes:
                # Cabeçalho das cargas
                for col_idx, col_name in enumerate(colunas_cargas):
                    cell = ws.cell(row=row, column=col_idx + 1, value=col_name)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center
                    cell.border = borda_fina
                ws.row_dimensions[row].height = 22
                row += 1
                
                cargas_count = 0
                total_peso = 0
                total_valor = 0
                
                # Fornecedores
                for f in detalhes.get('fornecedores', []):
                    peso = float(f.get('peso_ticket') or 0) if f.get('peso_ticket') else 0
                    valor = (f.get('valor_faturado') or 0) / 100
                    total_peso += peso
                    total_valor += valor
                    cargas_count += 1
                    
                    ws.cell(row=row, column=1, value=f.get('data_entrega', '-')).font = font_normal
                    ws.cell(row=row, column=2, value='Fornecedor').font = font_normal
                    ws.cell(row=row, column=3, value=(f.get('fornecedor_identificacao') or '-')[:30]).font = font_normal
                    ws.cell(row=row, column=4, value=(f.get('cliente') or '-')[:20]).font = font_normal
                    ws.cell(row=row, column=5, value=f.get('produto', '-')).font = font_normal
                    ws.cell(row=row, column=6, value=f.get('bitola', '-')).font = font_normal
                    ws.cell(row=row, column=7, value=f.get('nota_fiscal', '-')).font = font_normal
                    ws.cell(row=row, column=8, value=f.get('peso_ticket', '-')).font = font_normal
                    preco_cell = ws.cell(row=row, column=9, value=(f.get('preco_custo') or 0) / 100)
                    preco_cell.number_format = '#,##0.00'
                    preco_cell.font = font_normal
                    valor_cell = ws.cell(row=row, column=10, value=valor)
                    valor_cell.number_format = '#,##0.00'
                    valor_cell.font = font_valor
                    
                    for col in range(1, num_colunas + 1):
                        ws.cell(row=row, column=col).alignment = align_center
                        ws.cell(row=row, column=col).border = borda_fina
                    row += 1
                
                # Transportadoras
                for t in detalhes.get('transportadoras', []):
                    peso = float(t.get('peso_ticket') or 0) if t.get('peso_ticket') else 0
                    valor = (t.get('valor_faturado') or 0) / 100
                    total_peso += peso
                    total_valor += valor
                    cargas_count += 1
                    
                    ws.cell(row=row, column=1, value=t.get('data_entrega', '-')).font = font_normal
                    ws.cell(row=row, column=2, value='Transportadora').font = font_normal
                    ws.cell(row=row, column=3, value=(t.get('transportadora_identificacao') or t.get('nome') or '-')[:30]).font = font_normal
                    ws.cell(row=row, column=4, value=(t.get('cliente') or '-')[:20]).font = font_normal
                    ws.cell(row=row, column=5, value=t.get('produto', '-')).font = font_normal
                    ws.cell(row=row, column=6, value=t.get('bitola', '-')).font = font_normal
                    ws.cell(row=row, column=7, value=t.get('nota_fiscal', '-')).font = font_normal
                    ws.cell(row=row, column=8, value=t.get('peso_ticket', '-')).font = font_normal
                    preco_cell = ws.cell(row=row, column=9, value=(t.get('preco_custo') or 0) / 100)
                    preco_cell.number_format = '#,##0.00'
                    preco_cell.font = font_normal
                    valor_cell = ws.cell(row=row, column=10, value=valor)
                    valor_cell.number_format = '#,##0.00'
                    valor_cell.font = font_valor
                    
                    for col in range(1, num_colunas + 1):
                        ws.cell(row=row, column=col).alignment = align_center
                        ws.cell(row=row, column=col).border = borda_fina
                    row += 1
                
                # Extratores
                for e in detalhes.get('extratores', []):
                    peso = float(e.get('peso_ticket') or 0) if e.get('peso_ticket') else 0
                    valor = (e.get('valor_faturado') or 0) / 100
                    total_peso += peso
                    total_valor += valor
                    cargas_count += 1
                    
                    ws.cell(row=row, column=1, value=e.get('data_entrega', '-')).font = font_normal
                    ws.cell(row=row, column=2, value='Extrator').font = font_normal
                    ws.cell(row=row, column=3, value=(e.get('extrator_identificacao') or '-')[:30]).font = font_normal
                    ws.cell(row=row, column=4, value=(e.get('cliente') or '-')[:20]).font = font_normal
                    ws.cell(row=row, column=5, value=e.get('produto', '-')).font = font_normal
                    ws.cell(row=row, column=6, value=e.get('bitola', '-')).font = font_normal
                    ws.cell(row=row, column=7, value=e.get('nota_fiscal', '-')).font = font_normal
                    ws.cell(row=row, column=8, value=e.get('peso_ticket', '-')).font = font_normal
                    preco_cell = ws.cell(row=row, column=9, value=(e.get('preco_custo') or 0) / 100)
                    preco_cell.number_format = '#,##0.00'
                    preco_cell.font = font_normal
                    valor_cell = ws.cell(row=row, column=10, value=valor)
                    valor_cell.number_format = '#,##0.00'
                    valor_cell.font = font_valor
                    
                    for col in range(1, num_colunas + 1):
                        ws.cell(row=row, column=col).alignment = align_center
                        ws.cell(row=row, column=col).border = borda_fina
                    row += 1
                
                # Comissionados
                for c in detalhes.get('comissionados', []):
                    peso = float(c.get('peso_ticket') or 0) if c.get('peso_ticket') else 0
                    valor = (c.get('valor_faturado') or 0) / 100
                    total_peso += peso
                    total_valor += valor
                    cargas_count += 1
                    
                    ws.cell(row=row, column=1, value=c.get('data_entrega', '-')).font = font_normal
                    ws.cell(row=row, column=2, value='Comissionado').font = font_normal
                    ws.cell(row=row, column=3, value=(c.get('comissionado_identificacao') or '-')[:30]).font = font_normal
                    ws.cell(row=row, column=4, value=(c.get('cliente') or '-')[:20]).font = font_normal
                    ws.cell(row=row, column=5, value=c.get('produto', '-')).font = font_normal
                    ws.cell(row=row, column=6, value=c.get('bitola', '-')).font = font_normal
                    ws.cell(row=row, column=7, value=c.get('nota_fiscal', '-')).font = font_normal
                    ws.cell(row=row, column=8, value=c.get('peso_ticket', '-')).font = font_normal
                    preco_cell = ws.cell(row=row, column=9, value=(c.get('preco_custo') or 0) / 100)
                    preco_cell.number_format = '#,##0.00'
                    preco_cell.font = font_normal
                    valor_cell = ws.cell(row=row, column=10, value=valor)
                    valor_cell.number_format = '#,##0.00'
                    valor_cell.font = font_valor
                    
                    for col in range(1, num_colunas + 1):
                        ws.cell(row=row, column=col).alignment = align_center
                        ws.cell(row=row, column=col).border = borda_fina
                    row += 1
                
                # Cargas a Receber (Clientes) - AR
                for car in detalhes.get('cargas_a_receber', []):
                    peso = float(car.get('peso_ticket') or 0) if car.get('peso_ticket') else 0
                    valor = (car.get('valor_faturado') or 0) / 100
                    total_peso += peso
                    total_valor += valor
                    cargas_count += 1
                    
                    ws.cell(row=row, column=1, value=car.get('data_entrega', '-')).font = font_normal
                    ws.cell(row=row, column=2, value='Cliente').font = font_normal
                    ws.cell(row=row, column=3, value=(car.get('cliente') or '-')[:30]).font = font_normal
                    ws.cell(row=row, column=4, value='-').font = font_normal
                    ws.cell(row=row, column=5, value=car.get('produto', '-')).font = font_normal
                    ws.cell(row=row, column=6, value=car.get('bitola', '-')).font = font_normal
                    ws.cell(row=row, column=7, value=car.get('nota_fiscal', '-')).font = font_normal
                    ws.cell(row=row, column=8, value=car.get('peso_ticket', '-')).font = font_normal
                    preco_cell = ws.cell(row=row, column=9, value=(car.get('preco_custo') or 0) / 100)
                    preco_cell.number_format = '#,##0.00'
                    preco_cell.font = font_normal
                    valor_cell = ws.cell(row=row, column=10, value=valor)
                    valor_cell.number_format = '#,##0.00'
                    valor_cell.font = font_valor
                    
                    for col in range(1, num_colunas + 1):
                        ws.cell(row=row, column=col).alignment = align_center
                        ws.cell(row=row, column=col).border = borda_fina
                    row += 1
                
                # Linha de subtotal do grupo
                ws.cell(row=row, column=1, value=f'Subtotal ({cargas_count} cargas)').font = font_subtotal
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
                peso_cell = ws.cell(row=row, column=8, value=total_peso)
                peso_cell.font = font_subtotal
                peso_cell.number_format = '#,##0.00'
                ws.cell(row=row, column=9, value='').font = font_subtotal
                saldo_cell = ws.cell(row=row, column=10, value=saldo_pendente)
                saldo_cell.font = font_pendente
                saldo_cell.number_format = '#,##0.00'
                
                for col in range(1, num_colunas + 1):
                    ws.cell(row=row, column=col).fill = fill_subtotal
                    ws.cell(row=row, column=col).alignment = align_center
                    ws.cell(row=row, column=col).border = borda_grupo
                ws.cell(row=row, column=1).alignment = align_left
                ws.row_dimensions[row].height = 22
                row += 1
            
            elif grupo.get('agendamentos') and any(a.get('is_nf_complementar') for a in grupo.get('agendamentos', [])):
                # NF Complementar - exibir detalhes dos agendamentos
                # Cabeçalho
                colunas_nfc = ['Código', 'Tipo', 'Descrição', 'Data', 'Situação']
                for col_idx, col_name in enumerate(colunas_nfc):
                    cell = ws.cell(row=row, column=col_idx + 1, value=col_name)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center
                    cell.border = borda_fina
                # Coluna de valor
                cell = ws.cell(row=row, column=len(colunas_nfc) + 1, value='Valor')
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = borda_fina
                ws.row_dimensions[row].height = 22
                row += 1
                
                for item in grupo.get('agendamentos', []):
                    ws.cell(row=row, column=1, value=item.get('codigo_faturamento', '-')).font = font_normal
                    ws.cell(row=row, column=2, value=item.get('tipo_operacao', '-')).font = font_normal
                    ws.cell(row=row, column=3, value=(item.get('descricao') or '-')[:40]).font = font_normal
                    data_emissao = item.get('data_emissao')
                    ws.cell(row=row, column=4, value=data_emissao.strftime('%d/%m/%Y') if data_emissao else '-').font = font_normal
                    ws.cell(row=row, column=5, value=item.get('situacao', '-')).font = font_normal
                    valor_item = (item.get('saldo_100') or 0) / 100
                    valor_cell = ws.cell(row=row, column=6, value=valor_item)
                    valor_cell.number_format = '#,##0.00'
                    valor_cell.font = font_pendente
                    for col in range(1, 7):
                        ws.cell(row=row, column=col).alignment = align_center
                        ws.cell(row=row, column=col).border = borda_fina
                    row += 1
            
            else:
                # Lançamento avulso ou sem detalhes
                descricao = grupo.get('lancamento_descricao') or 'Detalhes não disponíveis'
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_colunas - 1)
                ws.cell(row=row, column=1, value=f'Lançamento: {descricao}').font = font_normal
                saldo_cell = ws.cell(row=row, column=num_colunas, value=saldo_pendente)
                saldo_cell.font = font_pendente
                saldo_cell.number_format = '#,##0.00'
                for col in range(1, num_colunas + 1):
                    ws.cell(row=row, column=col).border = borda_fina
                row += 1
            
            row += 1  # Linha em branco entre grupos
        
        # Linha de TOTAL GERAL
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_colunas - 1)
        total_cell = ws.cell(row=row, column=1, value='TOTAL GERAL - SALDO PENDENTE')
        total_cell.font = Font(name='Segoe UI', size=12, bold=True, color='FFFFFF')
        total_cell.fill = PatternFill(start_color=vermelho, end_color=vermelho, fill_type='solid')
        total_cell.alignment = align_center
        
        valor_total_cell = ws.cell(row=row, column=num_colunas, value=total_geral_saldo)
        valor_total_cell.font = Font(name='Segoe UI', size=12, bold=True, color='FFFFFF')
        valor_total_cell.fill = PatternFill(start_color=vermelho, end_color=vermelho, fill_type='solid')
        valor_total_cell.number_format = '#,##0.00'
        valor_total_cell.alignment = align_center
        ws.row_dimensions[row].height = 28
        
        # Ajustar largura das colunas
        larguras = [12, 14, 30, 20, 12, 12, 10, 12, 12, 14]
        for col_idx, largura in enumerate(larguras):
            col_letter = get_column_letter(col_idx + 1)
            ws.column_dimensions[col_letter].width = largura
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        resposta = make_response(output.read())
        resposta.headers['Content-Disposition'] = f'attachment; filename={nome_arquivo}.xlsx'
        resposta.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        return resposta

    @staticmethod
    def exportar_excel_agrupado_ap_pagamentos(grupos, nome_arquivo, titulo_planilha='AP - Pagamentos', direcao='ap'):
        """
        Gera um Excel com agrupamento visual por faturamento para AP Pagamentos ou AR Recebimentos.
        Similar ao de pendentes mas mostra valor pago ao invés de saldo pendente.
        Para AR, remove a coluna Cliente.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Dados'
        
        # Estilos
        verde_escuro = '1E5631'
        verde_claro = 'E8F5E8'
        azul_escuro = '1565C0'
        verde_pago = '2B8A3E'
        amarelo_parcial = 'D97706'
        cinza = 'F5F5F5'
        
        font_titulo = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
        fill_titulo = PatternFill(start_color=verde_escuro, end_color=verde_escuro, fill_type='solid')
        
        font_grupo = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        fill_grupo = PatternFill(start_color=verde_escuro, end_color=verde_escuro, fill_type='solid')
        
        font_header = Font(name='Segoe UI', size=10, bold=True, color=verde_escuro)
        fill_header = PatternFill(start_color=verde_claro, end_color=verde_claro, fill_type='solid')
        
        font_normal = Font(name='Segoe UI', size=10, color='333333')
        font_valor = Font(name='Segoe UI', size=10, bold=True, color=verde_escuro)
        font_pago = Font(name='Segoe UI', size=10, bold=True, color=verde_pago)
        font_parcial = Font(name='Segoe UI', size=10, bold=True, color=amarelo_parcial)
        
        font_subtotal = Font(name='Segoe UI', size=10, bold=True, color=verde_escuro)
        fill_subtotal = PatternFill(start_color=verde_claro, end_color=verde_claro, fill_type='solid')
        
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        
        borda_fina = Border(bottom=Side(style='thin', color='DDDDDD'))
        borda_grupo = Border(bottom=Side(style='medium', color=verde_escuro))
        
        # Colunas das cargas - AR não tem coluna Cliente
        if direcao == 'ar':
            colunas_cargas = ['Data', 'Tipo', 'Entidade', 'Produto', 'Bitola', 'NF', 'Peso (Ton)', 'Preço', 'Valor Final']
            col_produto = 4
            col_bitola = 5
            col_nf = 6
            col_peso = 7
            col_preco = 8
            col_valor = 9
        else:
            colunas_cargas = ['Data', 'Tipo', 'Entidade', 'Cliente', 'Produto', 'Bitola', 'NF', 'Peso (Ton)', 'Preço', 'Valor Final']
            col_produto = 5
            col_bitola = 6
            col_nf = 7
            col_peso = 8
            col_preco = 9
            col_valor = 10
        num_colunas = len(colunas_cargas)
        
        # Título do relatório (linha 1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_colunas)
        celula_titulo = ws.cell(row=1, column=1, value=titulo_planilha)
        celula_titulo.font = font_titulo
        celula_titulo.fill = fill_titulo
        celula_titulo.alignment = align_center
        ws.row_dimensions[1].height = 30
        
        row = 3  # Começa na linha 3
        
        total_geral_pago = 0
        
        for grupo in grupos:
            codigo_fat = grupo.get('codigo_faturamento', '-')
            pessoa_nome = grupo.get('pessoa_nome', '-')
            tipo_operacao = grupo.get('tipo_operacao', '-')
            data_pagamento = grupo.get('data_pagamento_mais_recente')
            total_pago = (grupo.get('total_pago_100') or 0) / 100
            data_faturamento = grupo.get('data_faturamento')
            tem_parcial = grupo.get('tem_parcial', False)
            
            total_geral_pago += total_pago
            
            # Status
            status = 'Pago Parcialmente' if tem_parcial else ('Pago' if direcao == 'ap' else 'Recebido')
            label_data = 'Pag' if direcao == 'ap' else 'Rec'
            
            # Linha do grupo (faturamento)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_colunas)
            info_grupo = f'{codigo_fat} - {pessoa_nome[:40]} | {tipo_operacao} | {status} | {label_data}: {data_pagamento.strftime("%d/%m/%Y") if data_pagamento else "-"} | Total: R$ {total_pago:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            celula_grupo = ws.cell(row=row, column=1, value=info_grupo)
            celula_grupo.font = font_grupo
            celula_grupo.fill = fill_grupo
            celula_grupo.alignment = align_left
            ws.row_dimensions[row].height = 25
            row += 1
            
            detalhes = grupo.get('detalhes_cargas')
            
            if detalhes:
                # Cabeçalho das cargas
                for col_idx, col_name in enumerate(colunas_cargas):
                    cell = ws.cell(row=row, column=col_idx + 1, value=col_name)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center
                    cell.border = borda_fina
                ws.row_dimensions[row].height = 22
                row += 1
                
                cargas_count = 0
                total_peso = 0
                total_valor = 0
                
                # Fornecedores
                for f in detalhes.get('fornecedores', []):
                    peso = float(f.get('peso_ticket') or 0) if f.get('peso_ticket') else 0
                    valor = (f.get('valor_faturado') or 0) / 100
                    total_peso += peso
                    total_valor += valor
                    cargas_count += 1
                    
                    ws.cell(row=row, column=1, value=f.get('data_entrega', '-')).font = font_normal
                    ws.cell(row=row, column=2, value='Fornecedor').font = font_normal
                    ws.cell(row=row, column=3, value=(f.get('fornecedor_identificacao') or '-')[:30]).font = font_normal
                    ws.cell(row=row, column=4, value=(f.get('cliente') or '-')[:20]).font = font_normal
                    ws.cell(row=row, column=5, value=f.get('produto', '-')).font = font_normal
                    ws.cell(row=row, column=6, value=f.get('bitola', '-')).font = font_normal
                    ws.cell(row=row, column=7, value=f.get('nota_fiscal', '-')).font = font_normal
                    ws.cell(row=row, column=8, value=f.get('peso_ticket', '-')).font = font_normal
                    preco_cell = ws.cell(row=row, column=9, value=(f.get('preco_custo') or 0) / 100)
                    preco_cell.number_format = '#,##0.00'
                    preco_cell.font = font_normal
                    valor_cell = ws.cell(row=row, column=10, value=valor)
                    valor_cell.number_format = '#,##0.00'
                    valor_cell.font = font_valor
                    
                    for col in range(1, num_colunas + 1):
                        ws.cell(row=row, column=col).alignment = align_center
                        ws.cell(row=row, column=col).border = borda_fina
                    row += 1
                
                # Transportadoras
                for t in detalhes.get('transportadoras', []):
                    peso = float(t.get('peso_ticket') or 0) if t.get('peso_ticket') else 0
                    valor = (t.get('valor_faturado') or 0) / 100
                    total_peso += peso
                    total_valor += valor
                    cargas_count += 1
                    
                    ws.cell(row=row, column=1, value=t.get('data_entrega', '-')).font = font_normal
                    ws.cell(row=row, column=2, value='Transportadora').font = font_normal
                    ws.cell(row=row, column=3, value=(t.get('transportadora_identificacao') or t.get('nome') or '-')[:30]).font = font_normal
                    ws.cell(row=row, column=4, value=(t.get('cliente') or '-')[:20]).font = font_normal
                    ws.cell(row=row, column=5, value=t.get('produto', '-')).font = font_normal
                    ws.cell(row=row, column=6, value=t.get('bitola', '-')).font = font_normal
                    ws.cell(row=row, column=7, value=t.get('nota_fiscal', '-')).font = font_normal
                    ws.cell(row=row, column=8, value=t.get('peso_ticket', '-')).font = font_normal
                    preco_cell = ws.cell(row=row, column=9, value=(t.get('preco_custo') or 0) / 100)
                    preco_cell.number_format = '#,##0.00'
                    preco_cell.font = font_normal
                    valor_cell = ws.cell(row=row, column=10, value=valor)
                    valor_cell.number_format = '#,##0.00'
                    valor_cell.font = font_valor
                    
                    for col in range(1, num_colunas + 1):
                        ws.cell(row=row, column=col).alignment = align_center
                        ws.cell(row=row, column=col).border = borda_fina
                    row += 1
                
                # Extratores
                for e in detalhes.get('extratores', []):
                    peso = float(e.get('peso_ticket') or 0) if e.get('peso_ticket') else 0
                    valor = (e.get('valor_faturado') or 0) / 100
                    total_peso += peso
                    total_valor += valor
                    cargas_count += 1
                    
                    ws.cell(row=row, column=1, value=e.get('data_entrega', '-')).font = font_normal
                    ws.cell(row=row, column=2, value='Extrator').font = font_normal
                    ws.cell(row=row, column=3, value=(e.get('extrator_identificacao') or '-')[:30]).font = font_normal
                    ws.cell(row=row, column=4, value=(e.get('cliente') or '-')[:20]).font = font_normal
                    ws.cell(row=row, column=5, value=e.get('produto', '-')).font = font_normal
                    ws.cell(row=row, column=6, value=e.get('bitola', '-')).font = font_normal
                    ws.cell(row=row, column=7, value=e.get('nota_fiscal', '-')).font = font_normal
                    ws.cell(row=row, column=8, value=e.get('peso_ticket', '-')).font = font_normal
                    preco_cell = ws.cell(row=row, column=9, value=(e.get('preco_custo') or 0) / 100)
                    preco_cell.number_format = '#,##0.00'
                    preco_cell.font = font_normal
                    valor_cell = ws.cell(row=row, column=10, value=valor)
                    valor_cell.number_format = '#,##0.00'
                    valor_cell.font = font_valor
                    
                    for col in range(1, num_colunas + 1):
                        ws.cell(row=row, column=col).alignment = align_center
                        ws.cell(row=row, column=col).border = borda_fina
                    row += 1
                
                # Comissionados
                for c in detalhes.get('comissionados', []):
                    peso = float(c.get('peso_ticket') or 0) if c.get('peso_ticket') else 0
                    valor = (c.get('valor_faturado') or 0) / 100
                    total_peso += peso
                    total_valor += valor
                    cargas_count += 1
                    
                    ws.cell(row=row, column=1, value=c.get('data_entrega', '-')).font = font_normal
                    ws.cell(row=row, column=2, value='Comissionado').font = font_normal
                    ws.cell(row=row, column=3, value=(c.get('comissionado_identificacao') or '-')[:30]).font = font_normal
                    ws.cell(row=row, column=4, value=(c.get('cliente') or '-')[:20]).font = font_normal
                    ws.cell(row=row, column=5, value=c.get('produto', '-')).font = font_normal
                    ws.cell(row=row, column=6, value=c.get('bitola', '-')).font = font_normal
                    ws.cell(row=row, column=7, value=c.get('nota_fiscal', '-')).font = font_normal
                    ws.cell(row=row, column=8, value=c.get('peso_ticket', '-')).font = font_normal
                    preco_cell = ws.cell(row=row, column=9, value=(c.get('preco_custo') or 0) / 100)
                    preco_cell.number_format = '#,##0.00'
                    preco_cell.font = font_normal
                    valor_cell = ws.cell(row=row, column=10, value=valor)
                    valor_cell.number_format = '#,##0.00'
                    valor_cell.font = font_valor
                    
                    for col in range(1, num_colunas + 1):
                        ws.cell(row=row, column=col).alignment = align_center
                        ws.cell(row=row, column=col).border = borda_fina
                    row += 1
                
                # Cargas a Receber (Clientes) - AR
                for car in detalhes.get('cargas_a_receber', []):
                    peso = float(car.get('peso_ticket') or 0) if car.get('peso_ticket') else 0
                    valor = (car.get('valor_faturado') or 0) / 100
                    total_peso += peso
                    total_valor += valor
                    cargas_count += 1
                    
                    ws.cell(row=row, column=1, value=car.get('data_entrega', '-')).font = font_normal
                    ws.cell(row=row, column=2, value='Cliente').font = font_normal
                    ws.cell(row=row, column=3, value=(car.get('cliente') or '-')[:30]).font = font_normal
                    # AR não tem coluna Cliente separada, usa índices dinâmicos
                    ws.cell(row=row, column=col_produto, value=car.get('produto', '-')).font = font_normal
                    ws.cell(row=row, column=col_bitola, value=car.get('bitola', '-')).font = font_normal
                    ws.cell(row=row, column=col_nf, value=car.get('nota_fiscal', '-')).font = font_normal
                    ws.cell(row=row, column=col_peso, value=car.get('peso_ticket', '-')).font = font_normal
                    preco_cell = ws.cell(row=row, column=col_preco, value=(car.get('preco_custo') or 0) / 100)
                    preco_cell.number_format = '#,##0.00'
                    preco_cell.font = font_normal
                    valor_cell = ws.cell(row=row, column=col_valor, value=valor)
                    valor_cell.number_format = '#,##0.00'
                    valor_cell.font = font_valor
                    
                    for col in range(1, num_colunas + 1):
                        ws.cell(row=row, column=col).alignment = align_center
                        ws.cell(row=row, column=col).border = borda_fina
                    row += 1
                
                # Linha de subtotal do grupo
                ws.cell(row=row, column=1, value=f'Subtotal ({cargas_count} cargas)').font = font_subtotal
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_nf)
                peso_cell = ws.cell(row=row, column=col_peso, value=total_peso)
                peso_cell.font = font_subtotal
                peso_cell.number_format = '#,##0.00'
                ws.cell(row=row, column=col_preco, value='').font = font_subtotal
                pago_cell = ws.cell(row=row, column=col_valor, value=total_pago)
                pago_cell.font = font_parcial if tem_parcial else font_pago
                pago_cell.number_format = '#,##0.00'
                
                for col in range(1, num_colunas + 1):
                    ws.cell(row=row, column=col).fill = fill_subtotal
                    ws.cell(row=row, column=col).alignment = align_center
                    ws.cell(row=row, column=col).border = borda_grupo
                ws.cell(row=row, column=1).alignment = align_left
                ws.row_dimensions[row].height = 22
                row += 1
            
            elif grupo.get('agendamentos') and any(a.get('is_nf_complementar') for a in grupo.get('agendamentos', [])):
                # NF Complementar - exibir detalhes dos agendamentos
                colunas_nfc = ['Código', 'Tipo', 'Descrição', 'Data', 'Situação']
                for col_idx, col_name in enumerate(colunas_nfc):
                    cell = ws.cell(row=row, column=col_idx + 1, value=col_name)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center
                    cell.border = borda_fina
                cell = ws.cell(row=row, column=len(colunas_nfc) + 1, value='Valor')
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = borda_fina
                ws.row_dimensions[row].height = 22
                row += 1
                
                for item in grupo.get('agendamentos', []):
                    ws.cell(row=row, column=1, value=item.get('codigo_faturamento', '-')).font = font_normal
                    ws.cell(row=row, column=2, value=item.get('tipo_operacao', '-')).font = font_normal
                    ws.cell(row=row, column=3, value=(item.get('descricao') or '-')[:40]).font = font_normal
                    data_emissao = item.get('data_emissao')
                    ws.cell(row=row, column=4, value=data_emissao.strftime('%d/%m/%Y') if data_emissao else '-').font = font_normal
                    ws.cell(row=row, column=5, value=item.get('situacao', '-')).font = font_normal
                    valor_item = (item.get('saldo_100') or 0) / 100
                    valor_cell = ws.cell(row=row, column=6, value=valor_item)
                    valor_cell.number_format = '#,##0.00'
                    valor_cell.font = font_pago
                    for col in range(1, 7):
                        ws.cell(row=row, column=col).alignment = align_center
                        ws.cell(row=row, column=col).border = borda_fina
                    row += 1
            
            else:
                # Lançamento avulso ou sem detalhes
                descricao = grupo.get('lancamento_descricao') or 'Detalhes não disponíveis'
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_colunas - 1)
                ws.cell(row=row, column=1, value=f'Lançamento: {descricao}').font = font_normal
                pago_cell = ws.cell(row=row, column=num_colunas, value=total_pago)
                pago_cell.font = font_parcial if tem_parcial else font_pago
                pago_cell.number_format = '#,##0.00'
                for col in range(1, num_colunas + 1):
                    ws.cell(row=row, column=col).border = borda_fina
                row += 1
            
            row += 1  # Linha em branco entre grupos
        
        # Linha de TOTAL GERAL
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_colunas - 1)
        label_total = 'TOTAL GERAL - VALOR PAGO' if direcao == 'ap' else 'TOTAL GERAL - VALOR RECEBIDO'
        total_cell = ws.cell(row=row, column=1, value=label_total)
        total_cell.font = Font(name='Segoe UI', size=12, bold=True, color='FFFFFF')
        total_cell.fill = PatternFill(start_color=verde_escuro, end_color=verde_escuro, fill_type='solid')
        total_cell.alignment = align_center
        
        valor_total_cell = ws.cell(row=row, column=num_colunas, value=total_geral_pago)
        valor_total_cell.font = Font(name='Segoe UI', size=12, bold=True, color='FFFFFF')
        valor_total_cell.fill = PatternFill(start_color=verde_escuro, end_color=verde_escuro, fill_type='solid')
        valor_total_cell.number_format = '#,##0.00'
        valor_total_cell.alignment = align_center
        ws.row_dimensions[row].height = 28
        
        # Ajustar largura das colunas
        if direcao == 'ar':
            larguras = [12, 14, 30, 12, 12, 10, 12, 12, 14]  # 9 colunas, sem Cliente
        else:
            larguras = [12, 14, 30, 20, 12, 12, 10, 12, 12, 14]  # 10 colunas
        for col_idx, largura in enumerate(larguras):
            col_letter = get_column_letter(col_idx + 1)
            ws.column_dimensions[col_letter].width = largura
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        resposta = make_response(output.read())
        resposta.headers['Content-Disposition'] = f'attachment; filename={nome_arquivo}.xlsx'
        resposta.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        return resposta