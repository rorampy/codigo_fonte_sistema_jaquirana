"""
Relatório de Recebimentos (AR) — Janeiro 2026
===================================================
Situações:
  2  = Pendente
  5  = Faturado
  6  = Categorizado
  7  = Não Categorizado
  8  = Conciliado
  9  = Liquidado
  10 = Parcialmente Conciliado

Fontes:
  Situações 2 e 5      → RegistroOperacionalModel (carga a receber)
  Situações 6,7,8,9,10 → AgendamentoPagamentoModel (via FaturamentoModel direcao_financeira=1)

Período: 01/01/2026 a 31/01/2026
Direção financeira = 1 (Receita / A Receber)
"""

import sys, os, json
from datetime import date
from decimal import Decimal

# Ajusta path para importar o sistema Flask
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))

from sistema import app, db

# ─── Models ──────────────────────────────────────────────────
from sistema.models_views.controle_carga.registro_operacional.registro_operacional_model import RegistroOperacionalModel
from sistema.models_views.controle_carga.solicitacao_nf.carga_model import CargaModel
from sistema.models_views.financeiro.operacional.categorizar_fatura.categorizacao_model import AgendamentoPagamentoModel
from sistema.models_views.financeiro.operacional.faturamento_model.faturamento_model import FaturamentoModel
from sistema.models_views.financeiro.lancamento_avulso.lancamento_avulso_model import LancamentoAvulsoModel
from sistema.models_views.faturamento.cargas_a_receber.nf_complementar.nf_complementar_model import NfComplementarModel
from sistema.models_views.gerenciar.pessoa_financeiro.pessoa_financeiro_model import PessoaFinanceiroModel
from sistema.models_views.configuracoes_gerais.plano_conta.plano_conta_model import PlanoContaModel

# ─── Constantes ──────────────────────────────────────────────
DATA_INICIO = date(2026, 1, 1)
DATA_FIM    = date(2026, 1, 31)

SITUACAO_PENDENTE              = 2
SITUACAO_FATURADO              = 5
SITUACAO_CATEGORIZADO          = 6
SITUACAO_NAO_CATEGORIZADO      = 7
SITUACAO_CONCILIADO            = 8
SITUACAO_LIQUIDADO             = 9
SITUACAO_PARCIALMENTE_CONCILIADO = 10

# Situações buscadas via RegistroOperacional (Pendente / Faturado)
SITUACOES_REGISTRO = [SITUACAO_PENDENTE, SITUACAO_FATURADO]

# Situações buscadas via AgendamentoPagamento (Categorizado, Não Cat., Conciliado, Liquidado, Parc. Conciliado)
SITUACOES_AGENDAMENTO = [SITUACAO_CATEGORIZADO, SITUACAO_NAO_CATEGORIZADO,
                         SITUACAO_CONCILIADO, SITUACAO_LIQUIDADO,
                         SITUACAO_PARCIALMENTE_CONCILIADO]

NOME_SITUACOES = {
    2:  'Pendente',
    5:  'Faturado',
    6:  'Categorizado',
    7:  'Não Categorizado',
    8:  'Conciliado',
    9:  'Liquidado',
    10: 'Parcialmente Conciliado',
}


# ═══════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════

def centavos_para_reais(valor_100):
    """Converte valor em centavos (int) para Decimal em reais."""
    if valor_100 is None:
        return Decimal('0.00')
    return Decimal(valor_100) / 100


def extrair_numero_nf(registro_oper):
    """Extrai o número da NF do registro operacional."""
    if not registro_oper:
        return ''
    if registro_oper.estorno_nf and registro_oper.numero_nota_fiscal_estorno:
        return f"{registro_oper.numero_nota_fiscal_estorno} *"
    return registro_oper.numero_nota_fiscal or ''


def extrair_plano_contas_json(categorias_json):
    """Extrai os nomes das categorias do plano de contas do JSON."""
    if not categorias_json:
        return ''
    try:
        categorias = json.loads(categorias_json) if isinstance(categorias_json, str) else categorias_json
        nomes = []
        for cat in (categorias or []):
            cat_id = cat.get('categoria_id')
            nome = cat.get('nome', '')
            if not nome and cat_id:
                plano = PlanoContaModel.query.get(cat_id)
                nome = plano.nome if plano else f'ID {cat_id}'
            if nome:
                nomes.append(nome)
        return ' | '.join(nomes)
    except Exception:
        return ''


def _data_entrega_no_periodo(data_entrega_str):
    """Verifica se a data_entrega (string dd/mm/yyyy) está dentro do período."""
    if not data_entrega_str:
        return False
    try:
        from datetime import datetime
        dt = datetime.strptime(data_entrega_str, '%d/%m/%Y').date()
        return DATA_INICIO <= dt <= DATA_FIM
    except (ValueError, TypeError):
        return False


# ═══════════════════════════════════════════════════════════════
#  PARTE 1 — SITUAÇÕES 2 e 5: RegistroOperacionalModel
#  (Pendente / Faturado — carga a receber)
# ═══════════════════════════════════════════════════════════════

def coletar_registros_operacionais():
    """
    Coleta registros de RegistroOperacionalModel nas situações 2 (Pendente) e 5 (Faturado).
    
    A data de referência é data_entrega_ticket.
    Direção financeira = 1 (Receita/A Receber) → identificada pelo campo situacao_financeira_id
    que referencia as mesmas situações de pagamento.
    """
    registros = RegistroOperacionalModel.query.filter(
        RegistroOperacionalModel.situacao_financeira_id.in_(SITUACOES_REGISTRO),
        RegistroOperacionalModel.ativo == True,
        RegistroOperacionalModel.deletado == False,
        RegistroOperacionalModel.data_entrega_ticket.isnot(None),
        RegistroOperacionalModel.data_entrega_ticket.between(DATA_INICIO, DATA_FIM),
    ).all()

    linhas = []

    for reg in registros:
        situacao_nome = NOME_SITUACOES.get(reg.situacao_financeira_id, str(reg.situacao_financeira_id))
        carga = reg.solicitacao  # CargaModel via solicitacao_nf_id

        # Dados da carga
        cliente = ''
        produto = ''
        bitola = ''
        fornecedor = ''

        if carga:
            cliente = carga.cliente.identificacao if carga.cliente else ''
            produto = carga.produto.nome if carga.produto else ''
            bitola = carga.bitola.bitola if carga.bitola else ''
            fornecedor = carga.fornecedor.identificacao if carga.fornecedor else ''

        # Número NF
        numero_nf = extrair_numero_nf(reg)

        # Peso
        peso_ton = reg.peso_liquido_ticket or ''

        # Preço unitário (em centavos)
        preco = centavos_para_reais(reg.preco_un_nf)

        # Valor total da nota (em centavos)
        valor_total = centavos_para_reais(reg.valor_total_nota_100)

        # Adiantamento = 0 para pendentes/faturados (ainda não conciliado)
        adiantamento = Decimal('0.00')

        # Líquido a Receber
        liquido = valor_total - adiantamento

        linhas.append({
            'Situação':          situacao_nome,
            'Tipo':              'Carga a Receber',
            'Data Emissão':      reg.data_entrega_ticket.strftime('%d/%m/%Y') if reg.data_entrega_ticket else '',
            'Número Documento':  numero_nf,
            'Entidade':          cliente,
            'Cliente':           cliente,
            'Produto':           produto,
            'Bitola':            bitola,
            'Peso Ton':          peso_ton,
            'Preço':             preco,
            'Valor Final':       valor_total,
            'Adiantamento':      adiantamento,
            'Líquido a Receber': liquido,
            'Plano de Contas':   '',
        })

    return linhas


# ═══════════════════════════════════════════════════════════════
#  PARTE 2 — SITUAÇÕES 6, 7, 8, 9, 10: AgendamentoPagamentoModel
#  (Categorizado, Não Categorizado, Conciliado, Liquidado, Parc. Conciliado)
# ═══════════════════════════════════════════════════════════════

def _nome_entidade_do_detalhe(detalhe, chave_grupo):
    """Extrai o nome da entidade do dict de detalhes de recebimento."""
    campos_id = {
        'cargas_a_receber': 'cliente',
        'nf_complementar':  'cliente',
        'nf_servico':       'cliente',
    }
    campo = campos_id.get(chave_grupo, 'cliente')
    return detalhe.get(campo, detalhe.get('cliente_identificacao', ''))


def _tipo_detalhe(chave_grupo):
    """Converte chave do grupo detalhes para tipo legível de recebimento."""
    mapa = {
        'cargas_a_receber': 'Carga a Receber',
        'nf_complementar':  'NF Complementar (Receita)',
        'nf_servico':       'NF Serviço',
    }
    return mapa.get(chave_grupo, chave_grupo)


def coletar_agendamentos_recebimento():
    """
    Coleta registros de AgendamentoPagamentoModel nas situações 6, 7, 8, 9 e 10
    vinculados a faturamentos com direcao_financeira = 1 (Receita).
    
    Também coleta lançamentos avulsos de receita (tipo_movimentacao = 1).
    """
    agendamentos = AgendamentoPagamentoModel.query.filter(
        AgendamentoPagamentoModel.situacao_pagamento_id.in_(SITUACOES_AGENDAMENTO),
        AgendamentoPagamentoModel.ativo == True,
        AgendamentoPagamentoModel.deletado == False,
        AgendamentoPagamentoModel.data_competencia.isnot(None),
        AgendamentoPagamentoModel.data_competencia.between(DATA_INICIO, DATA_FIM),
    ).all()

    linhas = []

    for ag in agendamentos:
        situacao_nome = NOME_SITUACOES.get(ag.situacao_pagamento_id, str(ag.situacao_pagamento_id))
        plano_contas = extrair_plano_contas_json(ag.categorias_json)

        # Determinar valor líquido conforme situação
        if ag.situacao_pagamento_id == SITUACAO_PARCIALMENTE_CONCILIADO:
            valor_conciliado = ag.valor_conciliado_100 or 0
            liquido_100 = (ag.valor_total_100 or 0) - valor_conciliado  # Falta receber
        else:
            valor_conciliado = 0
            liquido_100 = ag.valor_total_100 or 0

        # ── CARGAS FATURADAS (com detalhes expandidos) ──
        if ag.faturamento_id and ag.faturamento:
            fat = ag.faturamento

            # Somente receita (direcao_financeira = 1)
            if fat.direcao_financeira != 1:
                continue

            detalhes = fat.obter_detalhes()

            # Grupos de recebimento nos detalhes
            grupos_receita = ['cargas_a_receber', 'nf_complementar', 'nf_servico']

            tem_detalhes = False
            for grupo in grupos_receita:
                itens = detalhes.get(grupo, [])
                for item in itens:
                    tem_detalhes = True
                    tipo = _tipo_detalhe(grupo)
                    entidade = _nome_entidade_do_detalhe(item, grupo)

                    valor_bruto = item.get('valor_bruto', 0) or item.get('valor_faturado', 0) or item.get('valor_total', 0) or 0
                    valor_credito = item.get('valor_credito', 0) or 0
                    valor_faturado = item.get('valor_faturado', valor_bruto) or item.get('valor_total', valor_bruto) or 0

                    # Para parcialmente conciliado: proporcional ao conciliado
                    if ag.situacao_pagamento_id == SITUACAO_PARCIALMENTE_CONCILIADO:
                        total_fat = ag.valor_total_100 or 1
                        proporcao = valor_faturado / total_fat if total_fat else 0
                        liquido_item = round(liquido_100 * proporcao)
                    else:
                        liquido_item = valor_faturado - valor_credito

                    preco_raw = item.get('preco_custo', 0) or 0

                    linhas.append({
                        'Situação':          situacao_nome,
                        'Tipo':              tipo,
                        'Data Emissão':      item.get('data_entrega', item.get('data_emissao', ag.data_competencia.strftime('%d/%m/%Y') if ag.data_competencia else '')),
                        'Número Documento':  item.get('nota_fiscal', item.get('numero_nf', '')),
                        'Entidade':          entidade,
                        'Cliente':           item.get('cliente', ''),
                        'Produto':           item.get('produto', ''),
                        'Bitola':            item.get('bitola', ''),
                        'Peso Ton':          item.get('peso_ticket', ''),
                        'Preço':             centavos_para_reais(preco_raw),
                        'Valor Final':       centavos_para_reais(valor_faturado),
                        'Adiantamento':      centavos_para_reais(valor_credito),
                        'Líquido a Receber': centavos_para_reais(liquido_item),
                        'Plano de Contas':   plano_contas,
                    })

            # Se não teve detalhes expandidos, mostra linha resumida
            if not tem_detalhes:
                entidade = ag.pessoa_financeiro.identificacao if ag.pessoa_financeiro else ''
                linhas.append({
                    'Situação':          situacao_nome,
                    'Tipo':              'Faturamento (sem detalhe)',
                    'Data Emissão':      ag.data_competencia.strftime('%d/%m/%Y') if ag.data_competencia else '',
                    'Número Documento':  ag.referencia or fat.codigo_faturamento or '',
                    'Entidade':          entidade,
                    'Cliente':           '',
                    'Produto':           '',
                    'Bitola':            '',
                    'Peso Ton':          '',
                    'Preço':             Decimal('0.00'),
                    'Valor Final':       centavos_para_reais(ag.valor_total_100),
                    'Adiantamento':      centavos_para_reais(valor_conciliado),
                    'Líquido a Receber': centavos_para_reais(liquido_100),
                    'Plano de Contas':   plano_contas,
                })

        # ── LANÇAMENTOS AVULSOS (Receita) ──
        elif ag.lancamento_avulso_id and ag.lancamento_avulso:
            lanc = ag.lancamento_avulso

            # Apenas receitas (tipo_movimentacao = 1)
            if lanc.tipo_movimentacao != 1:
                continue

            entidade = ag.pessoa_financeiro.identificacao if ag.pessoa_financeiro else ''

            linhas.append({
                'Situação':          situacao_nome,
                'Tipo':              'Lançamento Avulso (Receita)',
                'Data Emissão':      ag.data_competencia.strftime('%d/%m/%Y') if ag.data_competencia else '',
                'Número Documento':  ag.referencia or '',
                'Entidade':          entidade,
                'Cliente':           '',
                'Produto':           '',
                'Bitola':            '',
                'Peso Ton':          '',
                'Preço':             Decimal('0.00'),
                'Valor Final':       centavos_para_reais(ag.valor_total_100),
                'Adiantamento':      centavos_para_reais(valor_conciliado),
                'Líquido a Receber': centavos_para_reais(liquido_100),
                'Plano de Contas':   plano_contas,
            })

        # ── AGENDAMENTO SEM FATURAMENTO E SEM LANÇAMENTO (caso genérico) ──
        else:
            entidade = ag.pessoa_financeiro.identificacao if ag.pessoa_financeiro else ''

            linhas.append({
                'Situação':          situacao_nome,
                'Tipo':              'Agendamento',
                'Data Emissão':      ag.data_competencia.strftime('%d/%m/%Y') if ag.data_competencia else '',
                'Número Documento':  ag.referencia or '',
                'Entidade':          entidade,
                'Cliente':           '',
                'Produto':           '',
                'Bitola':            '',
                'Peso Ton':          '',
                'Preço':             Decimal('0.00'),
                'Valor Final':       centavos_para_reais(ag.valor_total_100),
                'Adiantamento':      centavos_para_reais(valor_conciliado),
                'Líquido a Receber': centavos_para_reais(liquido_100),
                'Plano de Contas':   plano_contas,
            })

    return linhas


# ═══════════════════════════════════════════════════════════════
#  PARTE 3 — NF COMPLEMENTAR (receita: peso Ticket > peso NF)
# ═══════════════════════════════════════════════════════════════

def coletar_nf_complementar_receita():
    """
    Coleta NF Complementar pendentes de recebimento (receita).
    Quando peso_liquido_ticket > peso_ton_nf, o cliente deve à empresa.
    Inclui NF Complementar emitidas (fin_nf_complementar) + não emitidas (re_registro_operacional).
    """
    linhas = []

    # ── NF Complementar NÃO EMITIDAS (do RegistroOperacional) ──
    # Onde peso ticket > peso NF = receita pendente (cliente deve à empresa)
    registros_nao_emitidos = RegistroOperacionalModel.query.filter(
        RegistroOperacionalModel.ativo == True,
        RegistroOperacionalModel.deletado == False,
        RegistroOperacionalModel.solicitacao_nf_id.isnot(None),
        RegistroOperacionalModel.peso_ton_nf.isnot(None),
        RegistroOperacionalModel.peso_liquido_ticket.isnot(None),
        RegistroOperacionalModel.preco_un_nf > 0,
        RegistroOperacionalModel.peso_liquido_ticket > RegistroOperacionalModel.peso_ton_nf,
        db.or_(
            RegistroOperacionalModel.status_emissao_nf_complementar_id.is_(None),
            RegistroOperacionalModel.status_emissao_nf_complementar_id == 2
        ),
        RegistroOperacionalModel.data_entrega_ticket.isnot(None),
        RegistroOperacionalModel.data_entrega_ticket.between(DATA_INICIO, DATA_FIM),
    ).all()

    for r in registros_nao_emitidos:
        diferenca = (r.peso_liquido_ticket or 0) - (r.peso_ton_nf or 0)
        valor_receita_100 = round(diferenca * (r.preco_un_nf or 0))

        cliente = ''
        produto = ''
        bitola = ''
        if r.solicitacao:
            cliente = r.solicitacao.cliente.identificacao if r.solicitacao.cliente else ''
            produto = r.solicitacao.produto.nome if r.solicitacao.produto else ''
            bitola = r.solicitacao.bitola.bitola if r.solicitacao.bitola else ''

        linhas.append({
            'Situação':          'NF Compl. Pendente (não emitida)',
            'Tipo':              'NF Complementar (Receita)',
            'Data Emissão':      r.data_entrega_ticket.strftime('%d/%m/%Y') if r.data_entrega_ticket else '',
            'Número Documento':  r.numero_nota_fiscal or '',
            'Entidade':          cliente,
            'Cliente':           cliente,
            'Produto':           produto,
            'Bitola':            bitola,
            'Peso Ton':          round(diferenca, 4),
            'Preço':             centavos_para_reais(r.preco_un_nf),
            'Valor Final':       centavos_para_reais(valor_receita_100),
            'Adiantamento':      Decimal('0.00'),
            'Líquido a Receber': centavos_para_reais(valor_receita_100),
            'Plano de Contas':   '',
        })

    # ── NF Complementar EMITIDAS (fin_nf_complementar) com situação pendente ──
    nfs_emitidas = NfComplementarModel.query.filter(
        NfComplementarModel.ativo == True,
        NfComplementarModel.deletado == False,
        NfComplementarModel.situacao_financeira_id == SITUACAO_PENDENTE,
        NfComplementarModel.destinatario_data_emissao.isnot(None),
        NfComplementarModel.destinatario_data_emissao.between(DATA_INICIO, DATA_FIM),
    ).all()

    for nf in nfs_emitidas:
        cliente = nf.cliente.identificacao if nf.cliente else ''

        produto = ''
        bitola = ''
        if nf.nf_complementar_detalhes:
            try:
                detalhes = json.loads(nf.nf_complementar_detalhes) if isinstance(nf.nf_complementar_detalhes, str) else nf.nf_complementar_detalhes
                if isinstance(detalhes, list) and len(detalhes) > 0:
                    primeiro = detalhes[0]
                    produto = primeiro.get('produto', '')
                    bitola = primeiro.get('bitola', '')
            except Exception:
                pass

        linhas.append({
            'Situação':          'NF Compl. Pendente (emitida)',
            'Tipo':              'NF Complementar (Receita)',
            'Data Emissão':      nf.destinatario_data_emissao.strftime('%d/%m/%Y') if nf.destinatario_data_emissao else '',
            'Número Documento':  nf.numero_nota_fiscal or '',
            'Entidade':          cliente,
            'Cliente':           cliente,
            'Produto':           produto,
            'Bitola':            bitola,
            'Peso Ton':          nf.peso_ton_nf,
            'Preço':             centavos_para_reais(nf.preco_un_nf),
            'Valor Final':       centavos_para_reais(nf.valor_total_nota_100),
            'Adiantamento':      Decimal('0.00'),
            'Líquido a Receber': centavos_para_reais(nf.valor_total_nota_100),
            'Plano de Contas':   '',
        })

    return linhas


# ═══════════════════════════════════════════════════════════════
#  EXPORTAÇÃO EXCEL
# ═══════════════════════════════════════════════════════════════

def exportar_excel(todas_linhas, caminho_arquivo):
    """Exporta o resultado para Excel com formatação profissional, filtros e cabeçalho fixo."""
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    colunas = [
        'Situação', 'Tipo', 'Data Emissão', 'Número Documento',
        'Entidade', 'Cliente', 'Produto', 'Bitola', 'Peso Ton',
        'Preço', 'Valor Final', 'Adiantamento', 'Líquido a Receber',
        'Plano de Contas'
    ]

    df = pd.DataFrame(todas_linhas, columns=colunas)

    # Converter Data Emissão para formato de data
    df['Data Emissão'] = pd.to_datetime(df['Data Emissão'], format='mixed', errors='coerce')

    # Converter colunas numéricas para float
    for col in ['Preço', 'Valor Final', 'Adiantamento', 'Líquido a Receber']:
        df[col] = df[col].astype(float)

    # Ordenar por Situação, Tipo, Data Emissão
    df = df.sort_values(['Situação', 'Tipo', 'Data Emissão'], na_position='last')

    # Salvar inicial com pandas
    df.to_excel(caminho_arquivo, index=False, sheet_name='A Receber Jan-2026')

    # Formatação com openpyxl
    wb = load_workbook(caminho_arquivo)
    ws = wb.active

    # ── Congelar cabeçalho (freeze panes na linha 2) ──
    ws.freeze_panes = 'A2'

    # ── Auto-filtro em todas as colunas ──
    ultima_col_letter = get_column_letter(len(colunas))
    ws.auto_filter.ref = f'A1:{ultima_col_letter}{ws.max_row}'

    # ── Estilos ──
    header_fill = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')  # Verde escuro (receita)
    header_font = Font(name='Calibri', color='FFFFFF', bold=True, size=11)
    borda_fina = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    borda_header = Border(
        left=Side(style='thin', color='1B5E20'),
        right=Side(style='thin', color='1B5E20'),
        top=Side(style='thin', color='1B5E20'),
        bottom=Side(style='medium', color='1B5E20'),
    )
    fonte_dados = Font(name='Calibri', size=10)
    fill_cinza = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')  # Verde claro zebra
    fill_branco = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # ── Header ──
    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = borda_header
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ── Formato das colunas monetárias e de dados ──
    formato_moeda = '#,##0.00'
    colunas_moeda_idx = {10, 11, 12, 13}  # J, K, L, M (1-based)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = fonte_dados
            cell.border = borda_fina
            col_idx = cell.column

            if col_idx in colunas_moeda_idx:
                cell.number_format = formato_moeda
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col_idx == 3:  # Data Emissão
                cell.number_format = 'DD/MM/YYYY'
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx == 9:  # Peso Ton
                if cell.value:
                    try:
                        cell.value = float(cell.value)
                        cell.number_format = '#,##0.0000'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    except (ValueError, TypeError):
                        cell.alignment = Alignment(vertical='center')
            elif col_idx in (1, 2, 4, 8):  # Situação, Tipo, Nº Doc, Bitola
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(vertical='center')

    # ── Linhas alternadas (zebra) ──
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
        fill = fill_cinza if idx % 2 == 0 else fill_branco
        for cell in row:
            cell.fill = fill

    # ── Ajustar largura das colunas ──
    larguras_minimas = {
        1: 18,   # Situação
        2: 22,   # Tipo
        3: 14,   # Data Emissão
        4: 16,   # Número Documento
        5: 35,   # Entidade
        6: 30,   # Cliente
        7: 20,   # Produto
        8: 12,   # Bitola
        9: 14,   # Peso Ton
        10: 14,  # Preço
        11: 16,  # Valor Final
        12: 16,  # Adiantamento
        13: 18,  # Líquido a Receber
        14: 30,  # Plano de Contas
    }

    for col_cells in ws.columns:
        col_idx = col_cells[0].column
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_cells[0].value or ''))
        for cell in col_cells[1:101]:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        largura_conteudo = max_len + 3
        largura_min = larguras_minimas.get(col_idx, 12)
        ws.column_dimensions[col_letter].width = min(max(largura_conteudo, largura_min), 50)

    # ── Linha de totais ──
    ultima_dados = ws.max_row
    linha_total = ultima_dados + 2

    total_fill = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')
    total_font = Font(name='Calibri', color='FFFFFF', bold=True, size=11)
    total_border = Border(
        left=Side(style='thin', color='1B5E20'),
        right=Side(style='thin', color='1B5E20'),
        top=Side(style='medium', color='1B5E20'),
        bottom=Side(style='medium', color='1B5E20'),
    )

    ws.cell(row=linha_total, column=1, value='TOTAIS')
    ws.cell(row=linha_total, column=9, value=f'{ultima_dados - 1} registros')

    for col_idx in range(1, len(colunas) + 1):
        cell = ws.cell(row=linha_total, column=col_idx)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = total_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for col_idx, col_letter in [(11, 'K'), (12, 'L'), (13, 'M')]:  # Valor Final, Adiantamento, Líquido
        formula = f'=SUM({col_letter}2:{col_letter}{ultima_dados})'
        cell = ws.cell(row=linha_total, column=col_idx, value=formula)
        cell.number_format = formato_moeda
        cell.font = total_font
        cell.fill = total_fill
        cell.border = total_border
        cell.alignment = Alignment(horizontal='right', vertical='center')

    # ── Configurações de impressão ──
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = '1:1'

    wb.save(caminho_arquivo)
    print(f"\n✅ Arquivo salvo em: {caminho_arquivo}")
    print(f"   Total de registros: {len(df)}")

    # Resumo por situação
    print("\n📊 Resumo por Situação:")
    resumo = df.groupby('Situação').agg(
        Qtd=('Situação', 'count'),
        Valor_Final=('Valor Final', 'sum'),
        Liquido=('Líquido a Receber', 'sum')
    )
    for sit, row in resumo.iterrows():
        print(f"   {sit}: {int(row['Qtd'])} registros | Valor: R$ {row['Valor_Final']:,.2f} | Líquido: R$ {row['Liquido']:,.2f}")

    # Resumo por tipo
    print("\n📊 Resumo por Tipo:")
    resumo_tipo = df.groupby('Tipo').agg(
        Qtd=('Tipo', 'count'),
        Valor_Final=('Valor Final', 'sum'),
        Liquido=('Líquido a Receber', 'sum')
    )
    for tipo, row in resumo_tipo.iterrows():
        print(f"   {tipo}: {int(row['Qtd'])} registros | Valor: R$ {row['Valor_Final']:,.2f} | Líquido: R$ {row['Liquido']:,.2f}")


# ═══════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════

def main():
    print("=" * 70)
    print(" RELATÓRIO DE RECEBIMENTOS (AR) — JANEIRO 2026")
    print(" Situações: 2, 5, 6, 7, 8, 9, 10 | Direção Financeira = 1 (Receita)")
    print("=" * 70)

    todas_linhas = []

    # ── Parte 1: Pendentes e Faturados (Situações 2, 5 — RegistroOperacionalModel) ──
    print("\n🔍 Coletando PENDENTES e FATURADOS (Situações 2 e 5 — RegistroOperacionalModel)...")
    registros = coletar_registros_operacionais()
    print(f"   {len(registros)} registros")
    todas_linhas.extend(registros)

    # ── Parte 2: Agendamentos (Situações 6, 7, 8, 9, 10 — AgendamentoPagamentoModel) ──
    print("\n🔍 Coletando AGENDAMENTOS (Situações 6, 7, 8, 9, 10 — AgendamentoPagamentoModel)...")
    agendamentos = coletar_agendamentos_recebimento()
    print(f"   {len(agendamentos)} registros")
    todas_linhas.extend(agendamentos)

    # ── Parte 3: NF Complementar (receita: peso ticket > peso NF) ──
    print("\n🔍 Coletando NF COMPLEMENTAR (receita: peso ticket > peso NF)...")
    nf_compl = coletar_nf_complementar_receita()
    print(f"   {len(nf_compl)} registros")
    todas_linhas.extend(nf_compl)

    # ── Exportação ──
    if not todas_linhas:
        print("\n⚠️  Nenhum registro encontrado no período.")
        return

    caminho = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        '..',
        'relatorio_a_receber_janeiro_2026.xlsx'
    )
    caminho = os.path.abspath(caminho)

    exportar_excel(todas_linhas, caminho)


if __name__ == '__main__':
    with app.app_context():
        main()
