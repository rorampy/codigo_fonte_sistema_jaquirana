"""
Relatório de NF Complementar — Janeiro 2026
=============================================
Mostra TODAS as diferenças de peso (NF vs Ticket) do período,
classificando cada registro como:

  • POSITIVA (peso NF > peso Ticket) → Empresa PAGA ao cliente (Despesa)
  • NEGATIVA (peso Ticket > peso NF) → Empresa RECEBE do cliente (Receita)

Fontes:
  1) RegistroOperacionalModel — registros com peso NF e peso Ticket preenchidos
  2) NfComplementarModel      — NFs complementares já emitidas no período

Período: 01/01/2026 a 31/01/2026
"""

import sys, os
from datetime import date, datetime
from decimal import Decimal

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))

from sistema import app, db

# ─── Models ──────────────────────────────────────────────────
from sistema.models_views.controle_carga.registro_operacional.registro_operacional_model import RegistroOperacionalModel
from sistema.models_views.controle_carga.solicitacao_nf.carga_model import CargaModel
from sistema.models_views.controle_carga.produto.produto_model import ProdutoModel
from sistema.models_views.parametros.bitola.bitola_model import BitolaModel
from sistema.models_views.gerenciar.cliente.cliente_model import ClienteModel
from sistema.models_views.gerenciar.fornecedor.fornecedor_cadastro_model import FornecedorCadastroModel
from sistema.models_views.gerenciar.transportadora.transportadora_model import TransportadoraModel
from sistema.models_views.faturamento.cargas_a_receber.nf_complementar.nf_complementar_model import NfComplementarModel

# ─── Constantes ──────────────────────────────────────────────
DATA_INICIO = date(2026, 1, 1)
DATA_FIM    = date(2026, 1, 31)


def centavos_para_reais(valor_100):
    """Converte valor em centavos (int) para Decimal em reais."""
    if valor_100 is None:
        return Decimal('0.00')
    return Decimal(valor_100) / 100


# ═══════════════════════════════════════════════════════════════
#  COLETA — RegistroOperacionalModel (todas as diferenças)
# ═══════════════════════════════════════════════════════════════

def coletar_registros_operacionais():
    """
    Coleta TODOS os registros operacionais do período que possuem
    peso NF e peso Ticket preenchidos, independente do sinal da diferença.
    """
    registros = (
        db.session.query(RegistroOperacionalModel, ClienteModel)
        .join(CargaModel, RegistroOperacionalModel.solicitacao_nf_id == CargaModel.id)
        .join(ClienteModel, CargaModel.cliente_id == ClienteModel.id)
        .filter(
            RegistroOperacionalModel.ativo == True,
            RegistroOperacionalModel.deletado == False,
            RegistroOperacionalModel.solicitacao_nf_id.isnot(None),
            RegistroOperacionalModel.peso_ton_nf.isnot(None),
            RegistroOperacionalModel.peso_liquido_ticket.isnot(None),
            RegistroOperacionalModel.data_entrega_ticket.isnot(None),
            RegistroOperacionalModel.data_entrega_ticket.between(DATA_INICIO, DATA_FIM),
        )
        .all()
    )

    linhas = []

    for reg, cliente in registros:
        peso_nf = reg.peso_ton_nf or 0
        peso_ticket = reg.peso_liquido_ticket or 0
        diferenca = round(peso_nf - peso_ticket, 4)

        # Ignorar diferença zero
        if diferenca == 0:
            continue

        preco_un = reg.preco_un_nf or 0  # centavos
        valor_diferenca_100 = round(abs(diferenca) * preco_un)

        # Classificação
        if diferenca > 0:
            tipo_nf = 'POSITIVA'
            direcao = 'A Pagar (Despesa)'
        else:
            tipo_nf = 'NEGATIVA'
            direcao = 'A Receber (Receita)'

        # Nome do NF (estorno = *)
        numero_nf = ''
        if reg.estorno_nf and reg.numero_nota_fiscal_estorno:
            numero_nf = f"{reg.numero_nota_fiscal_estorno} *"
        else:
            numero_nf = reg.numero_nota_fiscal or ''

        # Status emissão NF complementar
        status_emissao = ''
        if reg.status_emissao_nf_complementar:
            status_emissao = reg.status_emissao_nf_complementar.status
        elif reg.status_emissao_nf_complementar_id is None:
            status_emissao = 'Não avaliado'
        else:
            status_emissao = 'Indefinido'

        # Dados da carga
        sol = reg.solicitacao
        produto = sol.produto.nome if sol and sol.produto else ''
        bitola = sol.bitola.bitola if sol and sol.bitola else ''

        # Fornecedor
        fornecedor = ''
        if sol and sol.fornecedor:
            fornecedor = sol.fornecedor.identificacao
        elif sol and sol.floresta:
            fornecedor = sol.floresta.identificacao

        # Transportadora
        transportadora = ''
        if sol and sol.transportadora_exibicao:
            transportadora = sol.transportadora_exibicao.identificacao

        linhas.append({
            'Data Entrega':      reg.data_entrega_ticket,
            'Data Emissão NF':   reg.destinatario_data_emissao,
            'Cliente':           cliente.identificacao,
            'Fornecedor':        fornecedor,
            'Transportadora':    transportadora,
            'Número NF':         numero_nf,
            'Produto':           produto,
            'Bitola':            bitola,
            'Placa':             reg.placa_nf or reg.placa_ticket or '',
            'Motorista':         reg.motorista_nf or reg.motorista_ticket or '',
            'Peso NF (Ton)':     round(peso_nf, 4),
            'Peso Ticket (Ton)': round(peso_ticket, 4),
            'Diferença (Ton)':   diferenca,
            'Preço Unitário':    centavos_para_reais(preco_un),
            'Valor Diferença':   centavos_para_reais(valor_diferenca_100),
            'Tipo NF Compl.':    tipo_nf,
            'Direção':           direcao,
            'Status Emissão':    status_emissao,
        })

    return linhas


# ═══════════════════════════════════════════════════════════════
#  COLETA — NfComplementarModel (NFs já emitidas)
# ═══════════════════════════════════════════════════════════════

def coletar_nf_complementares_emitidas():
    """
    Coleta NFs complementares já emitidas no período.
    Essas já possuem valor_total_nota_100 e peso consolidado.
    """
    nfs = NfComplementarModel.query.filter(
        NfComplementarModel.ativo == True,
        NfComplementarModel.deletado == False,
        NfComplementarModel.destinatario_data_emissao.isnot(None),
        NfComplementarModel.destinatario_data_emissao.between(DATA_INICIO, DATA_FIM),
    ).all()

    linhas = []

    for nf in nfs:
        cliente_nome = nf.cliente.identificacao if nf.cliente else ''
        valor_total = nf.valor_total_nota_100 or 0

        # Situação financeira
        situacao = ''
        if nf.situacao:
            situacao = nf.situacao.descricao if hasattr(nf.situacao, 'descricao') else str(nf.situacao_financeira_id)

        linhas.append({
            'NF Compl. Nº':       nf.numero_nota_fiscal or '',
            'Data Emissão':       nf.destinatario_data_emissao,
            'Cliente':            cliente_nome,
            'Destinatário':       nf.destinatario_nome or '',
            'CNPJ/CPF':           nf.destinatario_cnpj_cpf or '',
            'Peso NF (Ton)':      round(nf.peso_ton_nf, 4) if nf.peso_ton_nf else 0,
            'Preço Unitário':     centavos_para_reais(nf.preco_un_nf),
            'Valor Total':        centavos_para_reais(valor_total),
            'Série':              nf.serie_nota or '',
            'Chave Acesso':       nf.chave_acesso or '',
            'Transportador':      nf.transportador_nome or '',
            'Placa':              nf.placa_nf or '',
            'Motorista':          nf.motorista_nf or '',
            'Situação Financeira': situacao,
        })

    return linhas


# ═══════════════════════════════════════════════════════════════
#  EXPORTAÇÃO EXCEL
# ═══════════════════════════════════════════════════════════════

def exportar_excel(linhas_diferencas, linhas_emitidas, caminho_arquivo):
    """Exporta o resultado para Excel com 2 abas: Diferenças e NFs Emitidas."""
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── ABA 1: Diferenças de Peso ──
    colunas_diff = [
        'Data Entrega', 'Data Emissão NF', 'Cliente', 'Fornecedor', 'Transportadora',
        'Número NF', 'Produto', 'Bitola', 'Placa', 'Motorista',
        'Peso NF (Ton)', 'Peso Ticket (Ton)', 'Diferença (Ton)',
        'Preço Unitário', 'Valor Diferença',
        'Tipo NF Compl.', 'Direção', 'Status Emissão',
    ]
    df_diff = pd.DataFrame(linhas_diferencas, columns=colunas_diff)

    # Converter datas
    for col_data in ['Data Entrega', 'Data Emissão NF']:
        df_diff[col_data] = pd.to_datetime(df_diff[col_data], format='mixed', errors='coerce')

    # Converter colunas numéricas
    for col in ['Peso NF (Ton)', 'Peso Ticket (Ton)', 'Diferença (Ton)', 'Preço Unitário', 'Valor Diferença']:
        df_diff[col] = df_diff[col].astype(float)

    # Ordenar por Cliente, Data Entrega
    df_diff = df_diff.sort_values(['Cliente', 'Data Entrega'], na_position='last')

    # ── ABA 2: NFs Complementares Emitidas ──
    colunas_emit = [
        'NF Compl. Nº', 'Data Emissão', 'Cliente', 'Destinatário', 'CNPJ/CPF',
        'Peso NF (Ton)', 'Preço Unitário', 'Valor Total',
        'Série', 'Chave Acesso', 'Transportador', 'Placa', 'Motorista',
        'Situação Financeira',
    ]
    df_emit = pd.DataFrame(linhas_emitidas, columns=colunas_emit)

    if not df_emit.empty:
        df_emit['Data Emissão'] = pd.to_datetime(df_emit['Data Emissão'], format='mixed', errors='coerce')
        for col in ['Peso NF (Ton)', 'Preço Unitário', 'Valor Total']:
            df_emit[col] = df_emit[col].astype(float)
        df_emit = df_emit.sort_values(['Cliente', 'Data Emissão'], na_position='last')

    # ── Salvar com pandas (2 abas) ──
    with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
        df_diff.to_excel(writer, index=False, sheet_name='Diferenças de Peso')
        df_emit.to_excel(writer, index=False, sheet_name='NFs Emitidas')

    # ── Formatação com openpyxl ──
    wb = load_workbook(caminho_arquivo)

    # Estilos comuns
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(name='Calibri', color='FFFFFF', bold=True, size=11)
    borda_fina = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    borda_header = Border(
        left=Side(style='thin', color='1F4E79'),
        right=Side(style='thin', color='1F4E79'),
        top=Side(style='thin', color='1F4E79'),
        bottom=Side(style='medium', color='1F4E79'),
    )
    fonte_dados = Font(name='Calibri', size=10)
    fill_cinza = PatternFill(start_color='F2F6FA', end_color='F2F6FA', fill_type='solid')
    fill_branco = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    fill_verde = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    fill_vermelho = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    formato_moeda = '#,##0.00'
    formato_peso = '#,##0.0000'

    # ────────── FORMATAR ABA 1: Diferenças de Peso ──────────
    ws = wb['Diferenças de Peso']
    ws.freeze_panes = 'A2'
    ultima_col = get_column_letter(len(colunas_diff))
    ws.auto_filter.ref = f'A1:{ultima_col}{ws.max_row}'

    # Header
    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = borda_header
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Colunas monetárias/peso (índices 1-based)
    cols_peso = {11, 12, 13}     # Peso NF, Peso Ticket, Diferença
    cols_moeda = {14, 15}        # Preço Unit, Valor Diferença
    cols_data = {1, 2}           # Data Entrega, Data Emissão NF
    cols_centro = {6, 8, 9, 16, 17, 18}  # Nº NF, Bitola, Placa, Tipo, Direção, Status

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
        for cell in row:
            cell.font = fonte_dados
            cell.border = borda_fina
            col_idx = cell.column

            if col_idx in cols_peso:
                cell.number_format = formato_peso
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col_idx in cols_moeda:
                cell.number_format = formato_moeda
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col_idx in cols_data:
                cell.number_format = 'DD/MM/YYYY'
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx in cols_centro:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(vertical='center')

        # Colorir linha conforme tipo (POSITIVA = vermelho claro, NEGATIVA = verde claro)
        tipo_cell = row[15]  # coluna 16 (Tipo NF Compl.) = índice 15 no array
        if tipo_cell.value == 'POSITIVA':
            for cell in row:
                cell.fill = fill_vermelho
        elif tipo_cell.value == 'NEGATIVA':
            for cell in row:
                cell.fill = fill_verde

    # Larguras
    larguras_diff = {
        1: 14, 2: 14, 3: 30, 4: 30, 5: 25,
        6: 16, 7: 18, 8: 12, 9: 12, 10: 20,
        11: 16, 12: 16, 13: 16,
        14: 16, 15: 18,
        16: 16, 17: 22, 18: 18,
    }
    for col_idx, width in larguras_diff.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Linha de Resumo ──
    ultima_dados = ws.max_row
    linha_sep = ultima_dados + 2

    total_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    total_font = Font(name='Calibri', color='FFFFFF', bold=True, size=11)
    total_border = Border(
        left=Side(style='thin', color='1F4E79'),
        right=Side(style='thin', color='1F4E79'),
        top=Side(style='medium', color='1F4E79'),
        bottom=Side(style='medium', color='1F4E79'),
    )

    # Totais gerais
    ws.cell(row=linha_sep, column=1, value='TOTAIS')
    ws.cell(row=linha_sep, column=10, value=f'{ultima_dados - 1} registros')

    for col_idx in range(1, len(colunas_diff) + 1):
        cell = ws.cell(row=linha_sep, column=col_idx)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = total_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Soma Diferença (Ton) e Valor Diferença
    for col_idx, col_letter in [(13, 'M'), (15, 'O')]:
        formula = f'=SUM({col_letter}2:{col_letter}{ultima_dados})'
        cell = ws.cell(row=linha_sep, column=col_idx, value=formula)
        cell.number_format = formato_moeda if col_idx == 15 else formato_peso
        cell.font = total_font
        cell.fill = total_fill
        cell.border = total_border
        cell.alignment = Alignment(horizontal='right', vertical='center')

    # ── Resumos adicionais (A Pagar vs A Receber) ──
    linha_resumo = linha_sep + 2

    # Cabeçalho resumo
    fill_resumo_header = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    font_resumo_header = Font(name='Calibri', color='FFFFFF', bold=True, size=11)

    ws.cell(row=linha_resumo, column=1, value='RESUMO').font = Font(name='Calibri', bold=True, size=12)
    linha_resumo += 1

    headers_resumo = ['Direção', 'Qtd Registros', 'Diferença Total (Ton)', 'Valor Total (R$)']
    for i, h in enumerate(headers_resumo, 1):
        c = ws.cell(row=linha_resumo, column=i, value=h)
        c.fill = fill_resumo_header
        c.font = font_resumo_header
        c.border = total_border
        c.alignment = Alignment(horizontal='center', vertical='center')

    # Calcular resumos
    total_pagar_ton = Decimal('0')
    total_pagar_valor = Decimal('0')
    qtd_pagar = 0
    total_receber_ton = Decimal('0')
    total_receber_valor = Decimal('0')
    qtd_receber = 0

    for li in linhas_diferencas:
        dif = Decimal(str(li['Diferença (Ton)']))
        val = li['Valor Diferença']
        if li['Tipo NF Compl.'] == 'POSITIVA':
            total_pagar_ton += abs(dif)
            total_pagar_valor += val
            qtd_pagar += 1
        else:
            total_receber_ton += abs(dif)
            total_receber_valor += val
            qtd_receber += 1

    # A Pagar
    linha_resumo += 1
    ws.cell(row=linha_resumo, column=1, value='A Pagar (Despesa)').font = Font(name='Calibri', bold=True, size=10, color='CC0000')
    ws.cell(row=linha_resumo, column=2, value=qtd_pagar).alignment = Alignment(horizontal='center')
    c = ws.cell(row=linha_resumo, column=3, value=float(total_pagar_ton))
    c.number_format = formato_peso
    c.alignment = Alignment(horizontal='right')
    c = ws.cell(row=linha_resumo, column=4, value=float(total_pagar_valor))
    c.number_format = formato_moeda
    c.alignment = Alignment(horizontal='right')
    for i in range(1, 5):
        ws.cell(row=linha_resumo, column=i).fill = fill_vermelho
        ws.cell(row=linha_resumo, column=i).border = borda_fina

    # A Receber
    linha_resumo += 1
    ws.cell(row=linha_resumo, column=1, value='A Receber (Receita)').font = Font(name='Calibri', bold=True, size=10, color='006600')
    ws.cell(row=linha_resumo, column=2, value=qtd_receber).alignment = Alignment(horizontal='center')
    c = ws.cell(row=linha_resumo, column=3, value=float(total_receber_ton))
    c.number_format = formato_peso
    c.alignment = Alignment(horizontal='right')
    c = ws.cell(row=linha_resumo, column=4, value=float(total_receber_valor))
    c.number_format = formato_moeda
    c.alignment = Alignment(horizontal='right')
    for i in range(1, 5):
        ws.cell(row=linha_resumo, column=i).fill = fill_verde
        ws.cell(row=linha_resumo, column=i).border = borda_fina

    # Saldo líquido
    linha_resumo += 1
    saldo = total_receber_valor - total_pagar_valor
    saldo_ton = total_receber_ton - total_pagar_ton
    ws.cell(row=linha_resumo, column=1, value='SALDO LÍQUIDO').font = Font(name='Calibri', bold=True, size=11)
    ws.cell(row=linha_resumo, column=2, value=f'{qtd_pagar + qtd_receber} total').alignment = Alignment(horizontal='center')
    c = ws.cell(row=linha_resumo, column=3, value=float(saldo_ton))
    c.number_format = formato_peso
    c.alignment = Alignment(horizontal='right')
    c.font = Font(name='Calibri', bold=True, size=11)
    c = ws.cell(row=linha_resumo, column=4, value=float(saldo))
    c.number_format = formato_moeda
    c.alignment = Alignment(horizontal='right')
    c.font = Font(name='Calibri', bold=True, size=11, color='006600' if saldo >= 0 else 'CC0000')
    for i in range(1, 5):
        ws.cell(row=linha_resumo, column=i).border = Border(
            top=Side(style='medium'), bottom=Side(style='medium'),
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        )

    # Configurações de impressão
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = '1:1'

    # ────────── FORMATAR ABA 2: NFs Emitidas ──────────
    ws2 = wb['NFs Emitidas']
    ws2.freeze_panes = 'A2'

    if df_emit.empty:
        ws2.cell(row=2, column=1, value='Nenhuma NF Complementar emitida no período.')
    else:
        ultima_col2 = get_column_letter(len(colunas_emit))
        ws2.auto_filter.ref = f'A1:{ultima_col2}{ws2.max_row}'

        ws2.row_dimensions[1].height = 30
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = borda_header
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        cols_moeda2 = {7, 8}     # Preço Unit, Valor Total
        cols_peso2 = {6}         # Peso NF
        cols_data2 = {2}         # Data Emissão

        for row_idx, row in enumerate(ws2.iter_rows(min_row=2, max_row=ws2.max_row), start=0):
            fill = fill_cinza if row_idx % 2 == 0 else fill_branco
            for cell in row:
                cell.font = fonte_dados
                cell.border = borda_fina
                cell.fill = fill
                col_idx = cell.column

                if col_idx in cols_moeda2:
                    cell.number_format = formato_moeda
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col_idx in cols_peso2:
                    cell.number_format = formato_peso
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col_idx in cols_data2:
                    cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(vertical='center')

        # Larguras aba 2
        larguras_emit = {
            1: 16, 2: 14, 3: 30, 4: 30, 5: 18,
            6: 16, 7: 16, 8: 18, 9: 10, 10: 50,
            11: 25, 12: 12, 13: 20, 14: 20,
        }
        for col_idx, width in larguras_emit.items():
            ws2.column_dimensions[get_column_letter(col_idx)].width = width

        # Totais aba 2
        ultima_dados2 = ws2.max_row
        linha_total2 = ultima_dados2 + 2

        ws2.cell(row=linha_total2, column=1, value='TOTAIS')
        ws2.cell(row=linha_total2, column=5, value=f'{ultima_dados2 - 1} NFs emitidas')
        for col_idx in range(1, len(colunas_emit) + 1):
            c = ws2.cell(row=linha_total2, column=col_idx)
            c.fill = total_fill
            c.font = total_font
            c.border = total_border
            c.alignment = Alignment(horizontal='center', vertical='center')

        # Soma Valor Total
        formula2 = f'=SUM(H2:H{ultima_dados2})'
        c = ws2.cell(row=linha_total2, column=8, value=formula2)
        c.number_format = formato_moeda
        c.font = total_font
        c.fill = total_fill
        c.border = total_border
        c.alignment = Alignment(horizontal='right', vertical='center')

        ws2.page_setup.orientation = 'landscape'
        ws2.page_setup.fitToWidth = 1
        ws2.page_setup.fitToHeight = 0
        ws2.print_title_rows = '1:1'

    wb.save(caminho_arquivo)
    print(f"\n✅ Arquivo salvo em: {caminho_arquivo}")
    print(f"   Aba 'Diferenças de Peso': {len(df_diff)} registros")
    print(f"   Aba 'NFs Emitidas': {len(df_emit)} registros")


# ═══════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════

def main():
    print("=" * 70)
    print(" RELATÓRIO DE NF COMPLEMENTAR — JANEIRO 2026")
    print(" Diferenças de Peso (NF vs Ticket) + NFs Emitidas")
    print("=" * 70)

    # ── Parte 1: Diferenças de peso (todos os registros) ──
    print("\n🔍 Coletando diferenças de peso (RegistroOperacional)...")
    linhas_diferencas = coletar_registros_operacionais()
    print(f"   {len(linhas_diferencas)} registros com diferença")

    # Resumo
    positivas = [l for l in linhas_diferencas if l['Tipo NF Compl.'] == 'POSITIVA']
    negativas = [l for l in linhas_diferencas if l['Tipo NF Compl.'] == 'NEGATIVA']
    print(f"   → POSITIVAS (A Pagar): {len(positivas)} registros")
    print(f"   → NEGATIVAS (A Receber): {len(negativas)} registros")

    total_pagar = sum(l['Valor Diferença'] for l in positivas)
    total_receber = sum(l['Valor Diferença'] for l in negativas)
    print(f"   → Valor A Pagar:   R$ {total_pagar:,.2f}")
    print(f"   → Valor A Receber: R$ {total_receber:,.2f}")
    print(f"   → Saldo Líquido:   R$ {(total_receber - total_pagar):,.2f}")

    # ── Parte 2: NFs Complementares emitidas ──
    print("\n🔍 Coletando NFs Complementares emitidas...")
    linhas_emitidas = coletar_nf_complementares_emitidas()
    print(f"   {len(linhas_emitidas)} NFs emitidas no período")

    # ── Exportação ──
    if not linhas_diferencas and not linhas_emitidas:
        print("\n⚠️  Nenhum registro encontrado no período.")
        return

    caminho = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        '..',
        'relatorio_nf_complementar_janeiro_2026.xlsx'
    )
    caminho = os.path.abspath(caminho)

    exportar_excel(linhas_diferencas, linhas_emitidas, caminho)

    # Resumo por cliente
    if linhas_diferencas:
        print("\n📊 Resumo por Cliente:")
        from collections import defaultdict
        por_cliente = defaultdict(lambda: {'pagar': Decimal('0'), 'receber': Decimal('0'), 'qtd': 0})
        for l in linhas_diferencas:
            cli = l['Cliente']
            por_cliente[cli]['qtd'] += 1
            if l['Tipo NF Compl.'] == 'POSITIVA':
                por_cliente[cli]['pagar'] += l['Valor Diferença']
            else:
                por_cliente[cli]['receber'] += l['Valor Diferença']

        for cli in sorted(por_cliente.keys()):
            dados = por_cliente[cli]
            saldo = dados['receber'] - dados['pagar']
            sinal = '+' if saldo >= 0 else ''
            print(f"   {cli}: {dados['qtd']} registros | A Pagar: R$ {dados['pagar']:,.2f} | A Receber: R$ {dados['receber']:,.2f} | Saldo: {sinal}R$ {saldo:,.2f}")


if __name__ == '__main__':
    with app.app_context():
        main()
