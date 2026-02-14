"""
Relatório de Pendentes de Pagamento - Janeiro 2026
===================================================
Situações:
  2  = Pendente
  5  = Faturado
  6  = Categorizado
  7  = Não Categorizado
  8  = Conciliado
  10 = Parcialmente Conciliado

Fontes:
  Situação 2,8     → FretePagarModel, FornecedorPagarModel, ExtratorPagarModel, ComissionadoPagarModel
  Situações 5,6,7,8,10 → AgendamentoPagamentoModel (cargas faturadas + lançamentos avulsos)
  NF Complementar  → RegistroOperacionalModel (peso NF > peso ticket = despesa)

Período: 01/01/2026 a 31/01/2026
Direção financeira = 2 (Despesa / A Pagar)
"""

import sys, os, json
from datetime import date
from decimal import Decimal

# Ajusta path para importar o sistema Flask
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))

from sistema import app, db

# ─── Models ──────────────────────────────────────────────────
from sistema.models_views.faturamento.cargas_a_faturar.transportadora.frete_a_pagar_model import FretePagarModel
from sistema.models_views.faturamento.cargas_a_faturar.fornecedor.fornecedor_a_pagar_model import FornecedorPagarModel
from sistema.models_views.faturamento.cargas_a_faturar.extrator.extrator_a_pagar_model import ExtratorPagarModel
from sistema.models_views.faturamento.cargas_a_faturar.comissionado.comissionado_a_pagar_model import ComissionadoPagarModel
from sistema.models_views.controle_carga.registro_operacional.registro_operacional_model import RegistroOperacionalModel
from sistema.models_views.controle_carga.solicitacao_nf.carga_model import CargaModel
from sistema.models_views.financeiro.operacional.categorizar_fatura.categorizacao_model import AgendamentoPagamentoModel
from sistema.models_views.financeiro.operacional.faturamento_model.faturamento_model import FaturamentoModel
from sistema.models_views.financeiro.lancamento_avulso.lancamento_avulso_model import LancamentoAvulsoModel
from sistema.models_views.faturamento.cargas_a_receber.nf_complementar.nf_complementar_model import NfComplementarModel
from sistema.models_views.gerenciar.pessoa_financeiro.pessoa_financeiro_model import PessoaFinanceiroModel
from sistema.models_views.configuracoes_gerais.plano_conta.plano_conta_model import PlanoContaModel

# ─── Constantes ──────────────────────────────────────────────
DATA_INICIO = date(2026, 1, 1)
DATA_FIM    = date(2026, 1, 31)

SITUACAO_PENDENTE              = 2
SITUACAO_FATURADO              = 5
SITUACAO_CATEGORIZADO          = 6
SITUACAO_NAO_CATEGORIZADO      = 7
SITUACAO_CONCILIADO      = 8
SITUACAO_PARCIALMENTE_CONCILIADO = 10

SITUACOES_AGENDAMENTO = [SITUACAO_FATURADO, SITUACAO_CATEGORIZADO, SITUACAO_NAO_CATEGORIZADO, SITUACAO_PARCIALMENTE_CONCILIADO]

NOME_SITUACOES = {
    2:  'Pendente',
    5:  'Faturado',
    6:  'Categorizado',
    7:  'Não Categorizado',
    8:  'Conciliado',
    10: 'Parcialmente Conciliado',
}


def centavos_para_reais(valor_100):
    """Converte valor em centavos (int) para Decimal em reais."""
    if valor_100 is None:
        return Decimal('0.00')
    return Decimal(valor_100) / 100


def obter_registro_operacional(solicitacao_id):
    """Busca o RegistroOperacionalModel pelo solicitacao_nf_id."""
    if not solicitacao_id:
        return None
    return RegistroOperacionalModel.query.filter(
        RegistroOperacionalModel.solicitacao_nf_id == solicitacao_id,
        RegistroOperacionalModel.ativo == True,
        RegistroOperacionalModel.deletado == False,
    ).first()


def extrair_numero_nf(registro_oper):
    """Extrai o número da NF do registro operacional."""
    if not registro_oper:
        return ''
    if registro_oper.estorno_nf and registro_oper.numero_nota_fiscal_estorno:
        return f"{registro_oper.numero_nota_fiscal_estorno} *"
    return registro_oper.numero_nota_fiscal or ''


def extrair_plano_contas_json(categorias_json):
    """Extrai os nomes das categorias do plano de contas do JSON."""
    if not categorias_json:
        return ''
    try:
        categorias = json.loads(categorias_json) if isinstance(categorias_json, str) else categorias_json
        nomes = []
        for cat in (categorias or []):
            cat_id = cat.get('categoria_id')
            nome = cat.get('nome', '')
            if not nome and cat_id:
                plano = PlanoContaModel.query.get(cat_id)
                nome = plano.nome if plano else f'ID {cat_id}'
            if nome:
                nomes.append(nome)
        return ' | '.join(nomes)
    except Exception:
        return ''


# ═══════════════════════════════════════════════════════════════
#  PARTE 1 — SITUAÇÃO 2 (PENDENTE): dos modelos *_a_pagar
# ═══════════════════════════════════════════════════════════════

def coletar_pendentes_frete():
    """Frete a Pagar (Transportadora) — Situações 2 e 8."""
    registros = FretePagarModel.query.filter(
        FretePagarModel.situacao_pagamento_id.in_([SITUACAO_PENDENTE, SITUACAO_CONCILIADO]),
        FretePagarModel.ativo == True,
        FretePagarModel.deletado == False,
        FretePagarModel.data_entrega_ticket.isnot(None),
        FretePagarModel.data_entrega_ticket.between(DATA_INICIO, DATA_FIM),
    ).all()

    linhas = []
    for r in registros:
        registro_oper = obter_registro_operacional(r.solicitacao_id)
        adiantamento = (r.valor_credito_100 or 0) + (r.valor_saldo_debitado_100 or 0)
        liquido = (r.valor_total_a_pagar_100 or 0) - adiantamento

        # Nome da transportadora
        entidade = ''
        if r.transportadora:
            entidade = r.transportadora.identificacao
        elif r.solicitacao and r.solicitacao.transportadora_exibicao:
            entidade = r.solicitacao.transportadora_exibicao.identificacao

        linhas.append({
            'Código Faturamento': '',
            'Situação':          NOME_SITUACOES.get(r.situacao_pagamento_id, 'Pendente'),
            'Tipo':              'Frete (Transportadora)',
            'Data Emissão':      r.data_entrega_ticket,
            'Número Documento':  extrair_numero_nf(registro_oper),
            'Entidade':          entidade,
            'Cliente':           r.solicitacao.cliente.identificacao if r.solicitacao and r.solicitacao.cliente else '',
            'Produto':           r.solicitacao.produto.nome if r.solicitacao and r.solicitacao.produto else '',
            'Bitola':            r.bitola.bitola if r.bitola else '',
            'Peso Ton':          registro_oper.peso_liquido_ticket if registro_oper else None,
            'Preço':             centavos_para_reais(r.preco_custo_bitola_100),
            'Valor Final':       centavos_para_reais(r.valor_total_a_pagar_100),
            'Adiantamento':      centavos_para_reais(adiantamento),
            'Líquido a Pagar':   centavos_para_reais(liquido),
            'Valor Restante':    Decimal('0.00') if r.situacao_pagamento_id == SITUACAO_CONCILIADO else centavos_para_reais(liquido),
            'Plano de Contas':   '',
        })
    return linhas


def coletar_pendentes_fornecedor():
    """Fornecedor a Pagar — Situações 2 e 8."""
    registros = FornecedorPagarModel.query.filter(
        FornecedorPagarModel.situacao_pagamento_id.in_([SITUACAO_PENDENTE, SITUACAO_CONCILIADO]),
        FornecedorPagarModel.ativo == True,
        FornecedorPagarModel.deletado == False,
        FornecedorPagarModel.data_entrega_ticket.isnot(None),
        FornecedorPagarModel.data_entrega_ticket.between(DATA_INICIO, DATA_FIM),
    ).all()

    linhas = []
    for r in registros:
        registro_oper = obter_registro_operacional(r.solicitacao_id)
        adiantamento = (r.valor_credito_100 or 0) + (r.valor_saldo_debitado_100 or 0)
        liquido = (r.valor_total_a_pagar_100 or 0) - adiantamento

        linhas.append({
            'Código Faturamento': '',
            'Situação':          NOME_SITUACOES.get(r.situacao_pagamento_id, 'Pendente'),
            'Tipo':              'Fornecedor',
            'Data Emissão':      r.data_entrega_ticket,
            'Número Documento':  extrair_numero_nf(registro_oper),
            'Entidade':          r.fornecedor.identificacao if r.fornecedor else '',
            'Cliente':           r.solicitacao.cliente.identificacao if r.solicitacao and r.solicitacao.cliente else '',
            'Produto':           r.solicitacao.produto.nome if r.solicitacao and r.solicitacao.produto else '',
            'Bitola':            r.bitola.bitola if r.bitola else '',
            'Peso Ton':          registro_oper.peso_liquido_ticket if registro_oper else None,
            'Preço':             centavos_para_reais(r.preco_custo_bitola_100),
            'Valor Final':       centavos_para_reais(r.valor_total_a_pagar_100),
            'Adiantamento':      centavos_para_reais(adiantamento),
            'Líquido a Pagar':   centavos_para_reais(liquido),
            'Valor Restante':    Decimal('0.00') if r.situacao_pagamento_id == SITUACAO_CONCILIADO else centavos_para_reais(liquido),
            'Plano de Contas':   '',
        })
    return linhas


def coletar_pendentes_extrator():
    """Extrator a Pagar — Situações 2 e 8."""
    registros = ExtratorPagarModel.query.filter(
        ExtratorPagarModel.situacao_pagamento_id.in_([SITUACAO_PENDENTE, SITUACAO_CONCILIADO]),
        ExtratorPagarModel.ativo == True,
        ExtratorPagarModel.deletado == False,
        ExtratorPagarModel.data_entrega_ticket.isnot(None),
        ExtratorPagarModel.data_entrega_ticket.between(DATA_INICIO, DATA_FIM),
    ).all()

    linhas = []
    for r in registros:
        registro_oper = obter_registro_operacional(r.solicitacao_id)
        adiantamento = (r.valor_credito_100 or 0) + (r.valor_saldo_debitado_100 or 0)
        liquido = (r.valor_total_a_pagar_100 or 0) - adiantamento

        # Obter nome do extrator
        entidade = ''
        try:
            extrator = r.obter_extrator()
            if extrator:
                entidade = extrator.identificacao
            elif r.fornecedor:
                entidade = f"Extrator ({r.fornecedor.identificacao})"
        except Exception:
            entidade = r.fornecedor.identificacao if r.fornecedor else ''

        linhas.append({
            'Código Faturamento': '',
            'Situação':          NOME_SITUACOES.get(r.situacao_pagamento_id, 'Pendente'),
            'Tipo':              'Extrator',
            'Data Emissão':      r.data_entrega_ticket,
            'Número Documento':  extrair_numero_nf(registro_oper),
            'Entidade':          entidade,
            'Cliente':           r.solicitacao.cliente.identificacao if r.solicitacao and r.solicitacao.cliente else '',
            'Produto':           r.solicitacao.produto.nome if r.solicitacao and r.solicitacao.produto else '',
            'Bitola':            r.bitola.bitola if r.bitola else '',
            'Peso Ton':          registro_oper.peso_liquido_ticket if registro_oper else None,
            'Preço':             centavos_para_reais(r.preco_custo_bitola_100),
            'Valor Final':       centavos_para_reais(r.valor_total_a_pagar_100),
            'Adiantamento':      centavos_para_reais(adiantamento),
            'Líquido a Pagar':   centavos_para_reais(liquido),
            'Valor Restante':    Decimal('0.00') if r.situacao_pagamento_id == SITUACAO_CONCILIADO else centavos_para_reais(liquido),
            'Plano de Contas':   '',
        })
    return linhas


def coletar_pendentes_comissionado():
    """Comissionado a Pagar — Situações 2 e 8."""
    registros = ComissionadoPagarModel.query.filter(
        ComissionadoPagarModel.situacao_pagamento_id.in_([SITUACAO_PENDENTE, SITUACAO_CONCILIADO]),
        ComissionadoPagarModel.ativo == True,
        ComissionadoPagarModel.deletado == False,
        ComissionadoPagarModel.data_entrega_ticket.isnot(None),
        ComissionadoPagarModel.data_entrega_ticket.between(DATA_INICIO, DATA_FIM),
    ).all()

    linhas = []
    for r in registros:
        registro_oper = obter_registro_operacional(r.solicitacao_id)
        adiantamento = (r.valor_credito_100 or 0) + (r.valor_saldo_debitado_100 or 0)
        liquido = (r.valor_total_a_pagar_100 or 0) - adiantamento

        entidade = r.comissionado.identificacao if r.comissionado else ''

        linhas.append({
            'Código Faturamento': '',
            'Situação':          NOME_SITUACOES.get(r.situacao_pagamento_id, 'Pendente'),
            'Tipo':              'Comissionado',
            'Data Emissão':      r.data_entrega_ticket,
            'Número Documento':  extrair_numero_nf(registro_oper),
            'Entidade':          entidade,
            'Cliente':           r.solicitacao.cliente.identificacao if r.solicitacao and r.solicitacao.cliente else '',
            'Produto':           r.solicitacao.produto.nome if r.solicitacao and r.solicitacao.produto else '',
            'Bitola':            r.bitola.bitola if r.bitola else '',
            'Peso Ton':          registro_oper.peso_liquido_ticket if registro_oper else None,
            'Preço':             centavos_para_reais(r.preco_custo_bitola_100),
            'Valor Final':       centavos_para_reais(r.valor_total_a_pagar_100),
            'Adiantamento':      centavos_para_reais(adiantamento),
            'Líquido a Pagar':   centavos_para_reais(liquido),
            'Valor Restante':    Decimal('0.00') if r.situacao_pagamento_id == SITUACAO_CONCILIADO else centavos_para_reais(liquido),
            'Plano de Contas':   '',
        })
    return linhas


# ═══════════════════════════════════════════════════════════════
#  PARTE 2A — SITUAÇÕES 5 e 7: direto do FaturamentoModel
#  (Faturado / Não Categorizado — ainda NÃO têm AgendamentoPagamento)
# ═══════════════════════════════════════════════════════════════

def _tipo_entidade_do_detalhe(chave_grupo):
    """Converte chave do grupo detalhes para tipo legível."""
    mapa = {
        'fornecedores':    'Fornecedor',
        'transportadoras': 'Frete (Transportadora)',
        'extratores':      'Extrator',
        'comissionados':   'Comissionado',
        'nf_complementar': 'NF Complementar',
        'nf_servico':      'NF Serviço',
        'cargas_a_receber':'A Receber',
    }
    return mapa.get(chave_grupo, chave_grupo)


def _nome_entidade_do_detalhe(detalhe, chave_grupo):
    """Extrai o nome da entidade do dict de detalhes."""
    campos_id = {
        'fornecedores':    'fornecedor_identificacao',
        'transportadoras': 'transportadora_identificacao',
        'extratores':      'extrator_identificacao',
        'comissionados':   'comissionado_identificacao',
        'nf_complementar': 'cliente',
        'nf_servico':      'cliente',
        'cargas_a_receber':'cliente',
    }
    campo = campos_id.get(chave_grupo, 'identificacao')
    return detalhe.get(campo, '')


def _data_entrega_no_periodo(data_entrega_str):
    """Verifica se a data_entrega (string dd/mm/yyyy) está dentro do período."""
    if not data_entrega_str:
        return False
    try:
        from datetime import datetime
        dt = datetime.strptime(data_entrega_str, '%d/%m/%Y').date()
        return DATA_INICIO <= dt <= DATA_FIM
    except (ValueError, TypeError):
        return False


def _extrair_creditos_do_faturamento(detalhes):
    """Extrai créditos/adiantamentos do detalhes_json agrupados por tipo de entidade.
    
    Retorna dict: {
        'fornecedores':    {fornecedor_id: total_credito, ...},
        'transportadoras': {transportadora_id: total_credito, ...},
        'extratores':      {extrator_id: total_credito, ...},
    }
    e o total geral de créditos.
    """
    creditos = {
        'fornecedores': {},
        'transportadoras': {},
        'extratores': {},
    }
    total = 0

    for c in detalhes.get('credito_fornecedor', []):
        ent_id = c.get('fornecedor_id') or c.get('entidade_id')
        valor = c.get('valor', 0) or c.get('valor_credito_100', 0) or 0
        if ent_id and valor:
            creditos['fornecedores'][ent_id] = creditos['fornecedores'].get(ent_id, 0) + valor
            total += valor

    for c in detalhes.get('credito_transportadora', []):
        ent_id = c.get('transportadora_id') or c.get('entidade_id')
        valor = c.get('valor', 0) or c.get('valor_credito_100', 0) or 0
        if ent_id and valor:
            creditos['transportadoras'][ent_id] = creditos['transportadoras'].get(ent_id, 0) + valor
            total += valor

    for c in detalhes.get('credito_extrator', []):
        ent_id = c.get('extrator_id') or c.get('entidade_id')
        valor = c.get('valor', 0) or c.get('valor_credito_100', 0) or 0
        if ent_id and valor:
            creditos['extratores'][ent_id] = creditos['extratores'].get(ent_id, 0) + valor
            total += valor

    return creditos, total


# Mapa de grupo → campo de ID da entidade
_GRUPO_ID_MAP = {
    'fornecedores':    'fornecedor_id',
    'transportadoras': 'transportadora_id',
    'extratores':      'extrator_id',
}


def coletar_faturados_nao_categorizados():
    """
    Coleta registros de FaturamentoModel nas situações 5 (Faturado) e 7 (Não Categorizado).
    
    Esses registros ainda NÃO passaram por categorização, portanto NÃO têm 
    AgendamentoPagamentoModel. A data de referência é a data_entrega dentro 
    do detalhes_json do próprio FaturamentoModel.
    
    Exclui faturamentos que já possuem AgendamentoPagamentoModel ativo 
    (esses são capturados por coletar_agendamentos).
    """
    situacoes = [SITUACAO_FATURADO, SITUACAO_NAO_CATEGORIZADO, SITUACAO_CONCILIADO]

    # Subquery: IDs de faturamentos que já possuem agendamento ativo
    ids_com_agendamento = db.session.query(
        AgendamentoPagamentoModel.faturamento_id
    ).filter(
        AgendamentoPagamentoModel.faturamento_id.isnot(None),
        AgendamentoPagamentoModel.ativo == True,
        AgendamentoPagamentoModel.deletado == False,
    ).subquery()

    faturamentos = FaturamentoModel.query.filter(
        FaturamentoModel.situacao_pagamento_id.in_(situacoes),
        FaturamentoModel.direcao_financeira == 2,  # Despesa / A Pagar
        FaturamentoModel.ativo == True,
        FaturamentoModel.deletado == False,
        ~FaturamentoModel.id.in_(ids_com_agendamento),  # Excluir os que já têm agendamento
    ).all()

    linhas = []

    for fat in faturamentos:
        situacao_nome = NOME_SITUACOES.get(fat.situacao_pagamento_id, str(fat.situacao_pagamento_id))
        detalhes = fat.obter_detalhes()
        creditos_fat, total_cred_fat = _extrair_creditos_do_faturamento(detalhes)

        grupos_despesa = ['fornecedores', 'transportadoras', 'extratores', 'comissionados', 'nf_complementar']

        for grupo in grupos_despesa:
            itens = detalhes.get(grupo, [])
            for item in itens:
                # Filtrar pela data_entrega do item dentro do período
                if not _data_entrega_no_periodo(item.get('data_entrega', '')):
                    continue

                tipo = _tipo_entidade_do_detalhe(grupo)
                entidade = _nome_entidade_do_detalhe(item, grupo)

                valor_bruto = item.get('valor_bruto', 0) or item.get('valor_faturado', 0) or 0
                valor_credito_item = item.get('valor_credito', 0) or 0
                valor_faturado = item.get('valor_faturado', valor_bruto) or 0
                preco_raw = item.get('preco_custo', 0) or 0

                # Adiantamento: usar crédito do item OU do faturamento (credito_* arrays)
                adiantamento = valor_credito_item
                campo_id = _GRUPO_ID_MAP.get(grupo)
                if adiantamento == 0 and campo_id:
                    ent_id = item.get(campo_id)
                    if ent_id:
                        cred_fat_grupo = creditos_fat.get(grupo, {})
                        cred_total_entidade = cred_fat_grupo.get(ent_id, 0)
                        if cred_total_entidade > 0:
                            # Distribuir proporcionalmente entre itens da mesma entidade
                            total_bruto_entidade = sum(
                                (i.get('valor_bruto', 0) or i.get('valor_faturado', 0) or 0)
                                for i in itens if i.get(campo_id) == ent_id
                            ) or 1
                            adiantamento = round(cred_total_entidade * valor_bruto / total_bruto_entidade)

                liquido_item = valor_faturado - adiantamento

                linhas.append({
                    'Código Faturamento': fat.codigo_faturamento or '',
                    'Situação':          situacao_nome,
                    'Tipo':              tipo,
                    'Data Emissão':      item.get('data_entrega', ''),
                    'Número Documento':  item.get('nota_fiscal', ''),
                    'Entidade':          entidade,
                    'Cliente':           item.get('cliente', ''),
                    'Produto':           item.get('produto', ''),
                    'Bitola':            item.get('bitola', ''),
                    'Peso Ton':          item.get('peso_ticket', ''),
                    'Preço':             centavos_para_reais(preco_raw),
                    'Valor Final':       centavos_para_reais(valor_faturado),
                    'Adiantamento':      centavos_para_reais(adiantamento),
                    'Líquido a Pagar':   centavos_para_reais(liquido_item),
                    'Valor Restante':    Decimal('0.00') if fat.situacao_pagamento_id == SITUACAO_CONCILIADO else centavos_para_reais(liquido_item),
                    'Plano de Contas':   '',  # Ainda não categorizado
                })

    return linhas


# ═══════════════════════════════════════════════════════════════
#  PARTE 2A-II — LANÇAMENTOS AVULSOS DE DESPESA (Pendentes/Faturados/Não Categorizados)
#  Situações 2, 5, 7 — vão direto para AgendamentoPagamentoModel
#  (sem passar por FaturamentoModel)
# ═══════════════════════════════════════════════════════════════

def coletar_lancamentos_avulsos_despesa_pendentes():
    """
    Coleta lançamentos avulsos de despesa que estão nas situações 2 (Pendente),
    5 (Faturado) ou 7 (Não Categorizado) no AgendamentoPagamentoModel.
    
    Esses registros vêm de LancamentoAvulsoModel → AgendamentoPagamentoModel
    (sem passar por FaturamentoModel) e não são capturados pelas funções de
    pendentes de cargas nem pela função de agendamentos (que só pega 6 e 10).
    """
    situacoes = [SITUACAO_PENDENTE, SITUACAO_FATURADO, SITUACAO_NAO_CATEGORIZADO, SITUACAO_CONCILIADO]

    agendamentos = AgendamentoPagamentoModel.query.filter(
        AgendamentoPagamentoModel.lancamento_avulso_id.isnot(None),
        AgendamentoPagamentoModel.faturamento_id.is_(None),
        AgendamentoPagamentoModel.situacao_pagamento_id.in_(situacoes),
        AgendamentoPagamentoModel.ativo == True,
        AgendamentoPagamentoModel.deletado == False,
        db.or_(
            AgendamentoPagamentoModel.data_competencia.between(DATA_INICIO, DATA_FIM),
            db.and_(
                AgendamentoPagamentoModel.data_competencia.is_(None),
                AgendamentoPagamentoModel.data_vencimento.between(DATA_INICIO, DATA_FIM),
            ),
        ),
    ).all()

    linhas = []

    for ag in agendamentos:
        lanc = ag.lancamento_avulso
        if not lanc:
            continue

        # Apenas despesas (tipo_movimentacao = 2)
        if lanc.tipo_movimentacao != 2:
            continue

        situacao_nome = NOME_SITUACOES.get(ag.situacao_pagamento_id, str(ag.situacao_pagamento_id))
        plano_contas = extrair_plano_contas_json(ag.categorias_json)
        entidade = ag.pessoa_financeiro.identificacao if ag.pessoa_financeiro else ''

        # Determinar valores
        valor_total = ag.valor_total_100 or 0
        valor_conciliado_ag = ag.valor_conciliado_100 or 0

        # Adiantamento = somente créditos reais (lançamentos avulsos não possuem)
        adiantamento_ag = 0
        liquido_ag = valor_total

        # Valor restante (quanto falta conciliar)
        if ag.situacao_pagamento_id == SITUACAO_CONCILIADO:
            valor_restante = Decimal('0.00')
        elif ag.situacao_pagamento_id == SITUACAO_PARCIALMENTE_CONCILIADO:
            valor_restante = centavos_para_reais(valor_total - valor_conciliado_ag)
        else:
            valor_restante = centavos_para_reais(valor_total)

        # Data de referência: data_competencia ou data_vencimento
        data_ref = ''
        if ag.data_competencia:
            data_ref = ag.data_competencia.strftime('%d/%m/%Y')
        elif ag.data_vencimento:
            data_ref = ag.data_vencimento.strftime('%d/%m/%Y')

        linhas.append({
            'Código Faturamento': '',
            'Situação':          situacao_nome,
            'Tipo':              'Lançamento Avulso (Despesa)',
            'Data Emissão':      data_ref,
            'Número Documento':  ag.referencia or '',
            'Entidade':          entidade,
            'Cliente':           '',
            'Produto':           '',
            'Bitola':            '',
            'Peso Ton':          '',
            'Preço':             Decimal('0.00'),
            'Valor Final':       centavos_para_reais(valor_total),
            'Adiantamento':      centavos_para_reais(adiantamento_ag),
            'Líquido a Pagar':   centavos_para_reais(liquido_ag),
            'Valor Restante':    valor_restante,
            'Plano de Contas':   plano_contas,
        })

    return linhas


# ═══════════════════════════════════════════════════════════════
#  PARTE 2B — SITUAÇÕES 6, 7, 8 e 10: do AgendamentoPagamentoModel
#  (Categorizado / Conciliado / Parcialmente Conciliado — já têm Agendamento)
# ═══════════════════════════════════════════════════════════════

def coletar_agendamentos():
    """
    Coleta registros de AgendamentoPagamentoModel nas situações
    6 (Categorizado), 7 (Não Categorizado), 8 (Conciliado) e 10 (Parcialmente Conciliado).
    
    Esses registros já possuem AgendamentoPagamento e usam data_competencia como referência.
    """
    situacoes = [SITUACAO_NAO_CATEGORIZADO, SITUACAO_CATEGORIZADO, SITUACAO_CONCILIADO, SITUACAO_PARCIALMENTE_CONCILIADO]

    agendamentos = AgendamentoPagamentoModel.query.filter(
        AgendamentoPagamentoModel.situacao_pagamento_id.in_(situacoes),
        AgendamentoPagamentoModel.ativo == True,
        AgendamentoPagamentoModel.deletado == False,
        AgendamentoPagamentoModel.data_competencia.isnot(None),
        AgendamentoPagamentoModel.data_competencia.between(DATA_INICIO, DATA_FIM),
    ).all()

    linhas = []

    for ag in agendamentos:
        situacao_nome = NOME_SITUACOES.get(ag.situacao_pagamento_id, str(ag.situacao_pagamento_id))
        plano_contas = extrair_plano_contas_json(ag.categorias_json)

        # Determinar valores
        valor_total = ag.valor_total_100 or 0
        valor_conciliado_ag = ag.valor_conciliado_100 or 0

        # Adiantamento = somente créditos reais (não inclui valor já pago/conciliado)
        adiantamento_ag = 0

        # Líquido = valor total menos adiantamentos (neste nível, o próprio valor total)
        liquido_ag = valor_total

        # Valor restante (quanto falta conciliar)
        if ag.situacao_pagamento_id == SITUACAO_CONCILIADO:
            valor_restante = Decimal('0.00')
        elif ag.situacao_pagamento_id == SITUACAO_PARCIALMENTE_CONCILIADO:
            valor_restante = centavos_para_reais(valor_total - valor_conciliado_ag)
        else:
            valor_restante = centavos_para_reais(valor_total)

        # ── CARGAS FATURADAS (com detalhes expandidos) ──
        if ag.faturamento_id and ag.faturamento:
            fat = ag.faturamento

            # Verificar direção financeira = 2 (despesa)
            if fat.direcao_financeira != 2:
                continue

            detalhes = fat.obter_detalhes()
            creditos_fat, total_cred_fat = _extrair_creditos_do_faturamento(detalhes)

            # Grupos de entidades nos detalhes
            grupos_despesa = ['fornecedores', 'transportadoras', 'extratores', 'comissionados', 'nf_complementar']

            tem_detalhes = False
            for grupo in grupos_despesa:
                itens = detalhes.get(grupo, [])
                for item in itens:
                    tem_detalhes = True
                    tipo = _tipo_entidade_do_detalhe(grupo)
                    entidade = _nome_entidade_do_detalhe(item, grupo)

                    valor_bruto = item.get('valor_bruto', 0) or item.get('valor_faturado', 0) or 0
                    valor_credito_item = item.get('valor_credito', 0) or 0
                    valor_faturado = item.get('valor_faturado', valor_bruto) or 0
                    preco_raw = item.get('preco_custo', 0) or 0

                    # Adiantamento: per-item credit OU faturamento-level credit
                    adiantamento = valor_credito_item
                    campo_id = _GRUPO_ID_MAP.get(grupo)
                    if adiantamento == 0 and campo_id:
                        ent_id = item.get(campo_id)
                        if ent_id:
                            cred_fat_grupo = creditos_fat.get(grupo, {})
                            cred_total_entidade = cred_fat_grupo.get(ent_id, 0)
                            if cred_total_entidade > 0:
                                total_bruto_entidade = sum(
                                    (i.get('valor_bruto', 0) or i.get('valor_faturado', 0) or 0)
                                    for i in itens if i.get(campo_id) == ent_id
                                ) or 1
                                adiantamento = round(cred_total_entidade * valor_bruto / total_bruto_entidade)

                    # Líquido = bruto - adiantamento (valor real do pagamento líquido)
                    liquido_item = valor_faturado - adiantamento

                    # Valor Restante: quanto falta conciliar
                    if ag.situacao_pagamento_id == SITUACAO_CONCILIADO:
                        restante_item = 0
                    elif ag.situacao_pagamento_id == SITUACAO_PARCIALMENTE_CONCILIADO:
                        # Proporção deste item no total do agendamento
                        total_liq_fat = ag.valor_total_100 or 1
                        proporcao = liquido_item / total_liq_fat if total_liq_fat else 0
                        restante_100 = (ag.valor_total_100 or 0) - (ag.valor_conciliado_100 or 0)
                        restante_item = round(restante_100 * proporcao)
                    else:
                        restante_item = liquido_item

                    linhas.append({
                        'Código Faturamento': fat.codigo_faturamento or '',
                        'Situação':          situacao_nome,
                        'Tipo':              tipo,
                        'Data Emissão':      item.get('data_entrega', ag.data_competencia.strftime('%d/%m/%Y') if ag.data_competencia else ''),
                        'Número Documento':  item.get('nota_fiscal', ''),
                        'Entidade':          entidade,
                        'Cliente':           item.get('cliente', ''),
                        'Produto':           item.get('produto', ''),
                        'Bitola':            item.get('bitola', ''),
                        'Peso Ton':          item.get('peso_ticket', ''),
                        'Preço':             centavos_para_reais(preco_raw),
                        'Valor Final':       centavos_para_reais(valor_faturado),
                        'Adiantamento':      centavos_para_reais(adiantamento),
                        'Líquido a Pagar':   centavos_para_reais(liquido_item),
                        'Valor Restante':    centavos_para_reais(restante_item),
                        'Plano de Contas':   plano_contas,
                    })

            # Se não teve detalhes expandidos, mostra linha resumida
            if not tem_detalhes:
                entidade = ag.pessoa_financeiro.identificacao if ag.pessoa_financeiro else ''
                linhas.append({
                    'Código Faturamento': fat.codigo_faturamento or '',
                    'Situação':          situacao_nome,
                    'Tipo':              'Faturamento (sem detalhe)',
                    'Data Emissão':      ag.data_competencia.strftime('%d/%m/%Y') if ag.data_competencia else '',
                    'Número Documento':  ag.referencia or fat.codigo_faturamento or '',
                    'Entidade':          entidade,
                    'Cliente':           '',
                    'Produto':           '',
                    'Bitola':            '',
                    'Peso Ton':          '',
                    'Preço':             Decimal('0.00'),
                    'Valor Final':       centavos_para_reais(ag.valor_total_100),
                    'Adiantamento':      centavos_para_reais(adiantamento_ag),
                    'Líquido a Pagar':   centavos_para_reais(liquido_ag),
                    'Valor Restante':    valor_restante,
                    'Plano de Contas':   plano_contas,
                })

        # ── LANÇAMENTOS AVULSOS ──
        elif ag.lancamento_avulso_id and ag.lancamento_avulso:
            lanc = ag.lancamento_avulso

            # Apenas despesas (tipo_movimentacao = 2)
            if lanc.tipo_movimentacao != 2:
                continue

            entidade = ag.pessoa_financeiro.identificacao if ag.pessoa_financeiro else ''

            linhas.append({
                'Código Faturamento': '',
                'Situação':          situacao_nome,
                'Tipo':              'Lançamento Avulso (Despesa)',
                'Data Emissão':      ag.data_competencia.strftime('%d/%m/%Y') if ag.data_competencia else '',
                'Número Documento':  ag.referencia or '',
                'Entidade':          entidade,
                'Cliente':           '',
                'Produto':           '',
                'Bitola':            '',
                'Peso Ton':          '',
                'Preço':             Decimal('0.00'),
                'Valor Final':       centavos_para_reais(ag.valor_total_100),
                'Adiantamento':      centavos_para_reais(adiantamento_ag),
                'Líquido a Pagar':   centavos_para_reais(liquido_ag),
                'Valor Restante':    valor_restante,
                'Plano de Contas':   plano_contas,
            })

        # ── AGENDAMENTO SEM FATURAMENTO E SEM LANÇAMENTO (caso genérico) ──
        else:
            entidade = ag.pessoa_financeiro.identificacao if ag.pessoa_financeiro else ''

            linhas.append({
                'Código Faturamento': '',
                'Situação':          situacao_nome,
                'Tipo':              'Agendamento',
                'Data Emissão':      ag.data_competencia.strftime('%d/%m/%Y') if ag.data_competencia else '',
                'Número Documento':  ag.referencia or '',
                'Entidade':          entidade,
                'Cliente':           '',
                'Produto':           '',
                'Bitola':            '',
                'Peso Ton':          '',
                'Preço':             Decimal('0.00'),
                'Valor Final':       centavos_para_reais(ag.valor_total_100),
                'Adiantamento':      centavos_para_reais(adiantamento_ag),
                'Líquido a Pagar':   centavos_para_reais(liquido_ag),
                'Valor Restante':    valor_restante,
                'Plano de Contas':   plano_contas,
            })

    return linhas


# ═══════════════════════════════════════════════════════════════
#  PARTE 3 — NF COMPLEMENTAR (despesa: peso NF > peso Ticket)
# ═══════════════════════════════════════════════════════════════

def coletar_nf_complementar_despesa():
    """
    Coleta NF Complementar pendentes de pagamento (despesa).
    Quando peso_ton_nf > peso_liquido_ticket, a empresa deve ao cliente.
    Inclui NF Complementar emitidas (fin_nf_complementar) + não emitidas (re_registro_operacional).
    """
    linhas = []

    # ── NF Complementar NÃO EMITIDAS (do RegistroOperacional) ──
    # Onde peso NF > peso ticket = despesa pendente
    registros_nao_emitidos = RegistroOperacionalModel.query.filter(
        RegistroOperacionalModel.ativo == True,
        RegistroOperacionalModel.deletado == False,
        RegistroOperacionalModel.solicitacao_nf_id.isnot(None),
        RegistroOperacionalModel.peso_ton_nf.isnot(None),
        RegistroOperacionalModel.peso_liquido_ticket.isnot(None),
        RegistroOperacionalModel.preco_un_nf > 0,
        RegistroOperacionalModel.peso_ton_nf > RegistroOperacionalModel.peso_liquido_ticket,
        db.or_(
            RegistroOperacionalModel.status_emissao_nf_complementar_id.is_(None),
            RegistroOperacionalModel.status_emissao_nf_complementar_id == 2
        ),
        RegistroOperacionalModel.data_entrega_ticket.isnot(None),
        RegistroOperacionalModel.data_entrega_ticket.between(DATA_INICIO, DATA_FIM),
    ).all()

    for r in registros_nao_emitidos:
        diferenca = (r.peso_ton_nf or 0) - (r.peso_liquido_ticket or 0)
        valor_despesa_100 = round(diferenca * (r.preco_un_nf or 0))

        cliente = ''
        produto = ''
        bitola = ''
        if r.solicitacao:
            cliente = r.solicitacao.cliente.identificacao if r.solicitacao.cliente else ''
            produto = r.solicitacao.produto.nome if r.solicitacao.produto else ''
            bitola = r.solicitacao.bitola.bitola if r.solicitacao.bitola else ''

        linhas.append({
            'Código Faturamento': '',
            'Situação':          'NF Compl. Pendente (não emitida)',
            'Tipo':              'NF Complementar (Despesa)',
            'Data Emissão':      r.data_entrega_ticket,
            'Número Documento':  r.numero_nota_fiscal or '',
            'Entidade':          cliente,
            'Cliente':           cliente,
            'Produto':           produto,
            'Bitola':            bitola,
            'Peso Ton':          round(diferenca, 4),
            'Preço':             centavos_para_reais(r.preco_un_nf),
            'Valor Final':       centavos_para_reais(valor_despesa_100),
            'Adiantamento':      Decimal('0.00'),
            'Líquido a Pagar':   centavos_para_reais(valor_despesa_100),
            'Valor Restante':    Decimal('0.00'),
            'Plano de Contas':   '',
        })

    return linhas


# ═══════════════════════════════════════════════════════════════
#  EXPORTAÇÃO EXCEL
# ═══════════════════════════════════════════════════════════════

def exportar_excel(todas_linhas, caminho_arquivo):
    """Exporta o resultado para Excel com formatação profissional, filtros e cabeçalho fixo."""
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    colunas = [
        'Código Faturamento', 'Situação', 'Tipo', 'Data Emissão', 'Número Documento',
        'Entidade', 'Cliente', 'Produto', 'Bitola', 'Peso Ton',
        'Preço', 'Valor Final', 'Adiantamento', 'Líquido a Pagar', 'Valor Restante',
        'Plano de Contas'
    ]

    df = pd.DataFrame(todas_linhas, columns=colunas)

    # Converter Data Emissão para formato de data
    df['Data Emissão'] = pd.to_datetime(df['Data Emissão'], format='mixed', errors='coerce')

    # Converter colunas numéricas para float
    for col in ['Preço', 'Valor Final', 'Adiantamento', 'Líquido a Pagar', 'Valor Restante']:
        df[col] = df[col].astype(float)

    # Ordenar por Código Faturamento (agrupando cargas), Situação, Tipo, Data Emissão
    df = df.sort_values(['Código Faturamento', 'Situação', 'Tipo', 'Data Emissão'], na_position='last')

    # Salvar inicial com pandas
    df.to_excel(caminho_arquivo, index=False, sheet_name='A Pagar Jan-2026')

    # Formatação com openpyxl
    wb = load_workbook(caminho_arquivo)
    ws = wb.active

    # ── Congelar cabeçalho (freeze panes na linha 2) ──
    ws.freeze_panes = 'A2'

    # ── Auto-filtro em todas as colunas ──
    ultima_col_letter = get_column_letter(len(colunas))
    ws.auto_filter.ref = f'A1:{ultima_col_letter}{ws.max_row}'

    # ── Estilos ──
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(name='Calibri', color='FFFFFF', bold=True, size=11)
    borda_fina = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    borda_header = Border(
        left=Side(style='thin', color='1F4E79'),
        right=Side(style='thin', color='1F4E79'),
        top=Side(style='thin', color='1F4E79'),
        bottom=Side(style='medium', color='1F4E79'),
    )
    fonte_dados = Font(name='Calibri', size=10)
    fill_cinza = PatternFill(start_color='F2F6FA', end_color='F2F6FA', fill_type='solid')
    fill_branco = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # ── Header ──
    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = borda_header
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ── Formato das colunas monetárias e de dados ──
    formato_moeda = '#,##0.00'
    colunas_moeda_idx = {11, 12, 13, 14, 15}  # K=Preço, L=Valor Final, M=Adiantamento, N=Líquido, O=Valor Restante

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = fonte_dados
            cell.border = borda_fina
            col_idx = cell.column

            if col_idx in colunas_moeda_idx:
                cell.number_format = formato_moeda
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col_idx == 4:  # Data Emissão
                cell.number_format = 'DD/MM/YYYY'
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx == 10:  # Peso Ton
                if cell.value:
                    try:
                        cell.value = float(cell.value)
                        cell.number_format = '#,##0.0000'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    except (ValueError, TypeError):
                        cell.alignment = Alignment(vertical='center')
            elif col_idx in (1, 2, 3, 5, 9):  # Cód Fat, Situação, Tipo, Nº Doc, Bitola
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(vertical='center')

    # ── Linhas alternadas (zebra) ──
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
        fill = fill_cinza if idx % 2 == 0 else fill_branco
        for cell in row:
            cell.fill = fill

    # ── Ajustar largura das colunas ──
    larguras_minimas = {
        1: 18,   # Código Faturamento
        2: 18,   # Situação
        3: 22,   # Tipo
        4: 14,   # Data Emissão
        5: 16,   # Número Documento
        6: 35,   # Entidade
        7: 30,   # Cliente
        8: 20,   # Produto
        9: 12,   # Bitola
        10: 14,  # Peso Ton
        11: 14,  # Preço
        12: 16,  # Valor Final
        13: 16,  # Adiantamento
        14: 18,  # Líquido a Pagar
        15: 18,  # Valor Restante
        16: 30,  # Plano de Contas
    }

    for col_cells in ws.columns:
        col_idx = col_cells[0].column
        col_letter = get_column_letter(col_idx)
        # Calcular largura pelo conteúdo (amostra das primeiras 100 linhas)
        max_len = len(str(col_cells[0].value or ''))
        for cell in col_cells[1:101]:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        largura_conteudo = max_len + 3
        largura_min = larguras_minimas.get(col_idx, 12)
        ws.column_dimensions[col_letter].width = min(max(largura_conteudo, largura_min), 50)

    # ── Linha de totais ──
    ultima_dados = ws.max_row
    linha_total = ultima_dados + 2

    total_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    total_font = Font(name='Calibri', color='FFFFFF', bold=True, size=11)
    total_border = Border(
        left=Side(style='thin', color='1F4E79'),
        right=Side(style='thin', color='1F4E79'),
        top=Side(style='medium', color='1F4E79'),
        bottom=Side(style='medium', color='1F4E79'),
    )

    ws.cell(row=linha_total, column=1, value='TOTAIS')
    # Quantidade de registros
    ws.cell(row=linha_total, column=10, value=f'{ultima_dados - 1} registros')

    for col_idx in range(1, len(colunas) + 1):
        cell = ws.cell(row=linha_total, column=col_idx)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = total_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for col_idx, col_letter in [(12, 'L'), (13, 'M'), (14, 'N'), (15, 'O')]:  # Valor Final, Adiantamento, Líquido, Restante
        formula = f'=SUM({col_letter}2:{col_letter}{ultima_dados})'
        cell = ws.cell(row=linha_total, column=col_idx, value=formula)
        cell.number_format = formato_moeda
        cell.font = total_font
        cell.fill = total_fill
        cell.border = total_border
        cell.alignment = Alignment(horizontal='right', vertical='center')

    # ── Configurações de impressão ──
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = '1:1'  # Repetir cabeçalho em cada página impressa

    wb.save(caminho_arquivo)
    print(f"\n✅ Arquivo salvo em: {caminho_arquivo}")
    print(f"   Total de registros: {len(df)}")

    # Resumo por situação
    print("\n📊 Resumo por Situação:")
    resumo = df.groupby('Situação').agg(
        Qtd=('Situação', 'count'),
        Valor_Final=('Valor Final', 'sum'),
        Liquido=('Líquido a Pagar', 'sum')
    )
    for sit, row in resumo.iterrows():
        print(f"   {sit}: {int(row['Qtd'])} registros | Valor: R$ {row['Valor_Final']:,.2f} | Líquido: R$ {row['Liquido']:,.2f}")

    # Resumo por tipo
    print("\n📊 Resumo por Tipo:")
    resumo_tipo = df.groupby('Tipo').agg(
        Qtd=('Tipo', 'count'),
        Valor_Final=('Valor Final', 'sum'),
        Liquido=('Líquido a Pagar', 'sum')
    )
    for tipo, row in resumo_tipo.iterrows():
        print(f"   {tipo}: {int(row['Qtd'])} registros | Valor: R$ {row['Valor_Final']:,.2f} | Líquido: R$ {row['Liquido']:,.2f}")


# ═══════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════

def main():
    print("=" * 70)
    print(" RELATÓRIO DE PENDENTES DE PAGAMENTO — JANEIRO 2026")
    print(" Situações: 2, 5, 6, 7, 8, 10 | Direção Financeira = 2 (Despesa)")
    print("=" * 70)

    todas_linhas = []

    # ── Parte 1: Pendentes (Situação 2) ──
    print("\n🔍 Coletando PENDENTES (Situação 2)...")

    print("   → Frete a Pagar (Transportadora)...")
    frete = coletar_pendentes_frete()
    print(f"     {len(frete)} registros")
    todas_linhas.extend(frete)

    print("   → Fornecedor a Pagar...")
    fornecedor = coletar_pendentes_fornecedor()
    print(f"     {len(fornecedor)} registros")
    todas_linhas.extend(fornecedor)

    print("   → Extrator a Pagar...")
    extrator = coletar_pendentes_extrator()
    print(f"     {len(extrator)} registros")
    todas_linhas.extend(extrator)

    print("   → Comissionado a Pagar...")
    comissionado = coletar_pendentes_comissionado()
    print(f"     {len(comissionado)} registros")
    todas_linhas.extend(comissionado)

    # ── Parte 2A: Faturados / Não Categorizados (Situações 5, 7 — FaturamentoModel) ──
    print("\n🔍 Coletando FATURADOS e NÃO CATEGORIZADOS (Situações 5 e 7 — FaturamentoModel)...")
    fat_nao_cat = coletar_faturados_nao_categorizados()
    print(f"   {len(fat_nao_cat)} registros")
    todas_linhas.extend(fat_nao_cat)

    # ── Parte 2A-II: Lançamentos Avulsos de Despesa (Situações 2, 5, 7 — AgendamentoPagamentoModel) ──
    print("\n🔍 Coletando LANÇAMENTOS AVULSOS DE DESPESA (Situações 2, 5, 7 — AgendamentoPagamentoModel)...")
    lanc_avulsos_pend = coletar_lancamentos_avulsos_despesa_pendentes()
    print(f"   {len(lanc_avulsos_pend)} registros")
    todas_linhas.extend(lanc_avulsos_pend)

    # ── Parte 2B: Agendamentos Categorizados / Conciliados / Parc. Conciliados (Situações 6, 7, 8, 10 — AgendamentoPagamentoModel) ──
    print("\n🔍 Coletando AGENDAMENTOS (Situações 6, 7, 8 e 10 — AgendamentoPagamentoModel)...")
    agendamentos = coletar_agendamentos()
    print(f"   {len(agendamentos)} registros")
    todas_linhas.extend(agendamentos)

    # ── Parte 3: NF Complementar (despesa) ──
    print("\n🔍 Coletando NF COMPLEMENTAR (despesa: peso NF > peso ticket)...")
    nf_compl = coletar_nf_complementar_despesa()
    print(f"   {len(nf_compl)} registros")
    todas_linhas.extend(nf_compl)

    # ── Exportação ──
    if not todas_linhas:
        print("\n⚠️  Nenhum registro encontrado no período.")
        return

    caminho = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        '..',
        'relatorio_a_pagar_janeiro_2026.xlsx'
    )
    caminho = os.path.abspath(caminho)

    exportar_excel(todas_linhas, caminho)


if __name__ == '__main__':
    with app.app_context():
        main()
